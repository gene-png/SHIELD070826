"""Filename slugifier + XLSX + PDF render smokes."""

from __future__ import annotations

import io
import uuid
from datetime import date

import pytest
from app.models.capability import (
    CapabilityDisposition,
    CapabilityItem,
    CapabilityList,
    CapabilityListStatus,
)
from app.tech_debt.exporters import build_context, render_pdf, render_xlsx
from app.tech_debt.filename import (
    SERVICE_SLUG_TECH_DEBT,
    deliverable_filename,
    mmddyy,
    slugify,
)

# ---------------------------------------------------------------------------
# Slugifier
# ---------------------------------------------------------------------------


@pytest.mark.unit
@pytest.mark.parametrize(
    "raw,expected",
    [
        ("Acme, Inc.", "Acme_Inc"),
        ("  Atlas  Defense   Solutions  ", "Atlas_Defense_Solutions"),
        ("MITRE ATT&CK Coverage", "MITRE_ATTCK_Coverage"),
        ("___leading", "leading"),
        ("trailing___", "trailing"),
        ("", "Unknown"),
        ("   ", "Unknown"),
        ("!!!", "Unknown"),
        ("Nexus Federal Solutions Inc.", "Nexus_Federal_Solutions_Inc"),
        ("KEEP_CASE_AS_ENTERED", "KEEP_CASE_AS_ENTERED"),
        ("CamelCaseName", "CamelCaseName"),
    ],
)
def test_slugify(raw: str, expected: str) -> None:
    assert slugify(raw) == expected


@pytest.mark.unit
def test_slugify_none_returns_unknown() -> None:
    assert slugify(None) == "Unknown"


@pytest.mark.unit
def test_mmddyy_format() -> None:
    assert mmddyy(date(2026, 5, 18)) == "051826"
    assert mmddyy(date(2026, 1, 3)) == "010326"


@pytest.mark.unit
def test_deliverable_filename_matches_spec_example() -> None:
    # Spec §15.5: "Nexus Federal Solutions Inc. + Tech Debt Review + 2026-05-18"
    # -> "Nexus_Federal_Solutions_Inc_Tech_Debt_Review051826.pdf"
    name = deliverable_filename(
        company="Nexus Federal Solutions Inc.",
        service_slug=SERVICE_SLUG_TECH_DEBT,
        extension="pdf",
        day=date(2026, 5, 18),
    )
    assert name == "Nexus_Federal_Solutions_Inc_Tech_Debt_Review051826.pdf"


@pytest.mark.unit
def test_deliverable_filename_v2_re_release() -> None:
    name = deliverable_filename(
        company="Atlas Defense Solutions",
        service_slug=SERVICE_SLUG_TECH_DEBT,
        extension="xlsx",
        day=date(2026, 5, 18),
        version=2,
    )
    assert name == "Atlas_Defense_Solutions_Tech_Debt_Review051826_v2.xlsx"


@pytest.mark.unit
def test_deliverable_filename_working_prefix() -> None:
    name = deliverable_filename(
        company="X",
        service_slug=SERVICE_SLUG_TECH_DEBT,
        extension="xlsx",
        day=date(2026, 5, 18),
        working=True,
    )
    assert name == "WORKING_X_Tech_Debt_Review051826.xlsx"


# ---------------------------------------------------------------------------
# Render smokes
# ---------------------------------------------------------------------------


def _item(**kwargs) -> CapabilityItem:
    defaults = {
        "id": uuid.uuid4(),
        "capability_list_id": uuid.uuid4(),
        "name": "Wiz",
        "vendor": "Wiz, Inc.",
        "category": "CNAPP",
        "function": "Cloud posture",
        "annual_cost_usd": 350_000,
        "license_count": 200,
        "notes": None,
        "confidence_pct": 92,
        "source_artifact_id": None,
        "disposition": None,
        "disposition_rationale": None,
        "consolidation_target_id": None,
    }
    defaults.update(kwargs)
    return CapabilityItem(**defaults)


@pytest.fixture()
def context_with_items():
    cap_list = CapabilityList(
        id=uuid.uuid4(),
        service_id=uuid.uuid4(),
        version=1,
        status=CapabilityListStatus.APPROVED,
    )
    items = [
        _item(name="Wiz", annual_cost_usd=350_000, disposition=CapabilityDisposition.KEEP),
        _item(
            name="Lacework",
            annual_cost_usd=120_000,
            disposition=CapabilityDisposition.CUT,
        ),
        _item(
            name="Splunk",
            category="SIEM",
            annual_cost_usd=480_000,
            disposition=CapabilityDisposition.CONSOLIDATE,
        ),
    ]
    return build_context(
        client_legal_name="Atlas Defense Solutions",
        service_title="Technical Debt Review",
        cap_list=cap_list,
        items=items,
    )


@pytest.mark.unit
def test_xlsx_render_produces_valid_workbook(context_with_items) -> None:
    from openpyxl import load_workbook

    raw = render_xlsx(context_with_items)
    assert isinstance(raw, bytes)
    assert len(raw) > 1024  # non-trivial

    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    assert ws.title == "Inventory"
    # Header row.
    assert ws.cell(row=1, column=1).value == "Name"
    assert ws.cell(row=1, column=5).value == "Annual Cost (USD)"
    # Data rows.
    row_values = [ws.cell(row=2, column=c).value for c in range(1, 11)]
    assert row_values[0] == "Wiz"
    assert row_values[4] == 350_000
    # Summary row total cost = 350k + 120k + 480k = 950k.
    found_total = False
    for row in ws.iter_rows(min_row=4):
        if row[0].value == "Total annual cost":
            assert row[4].value == 950_000
            found_total = True
    assert found_total


@pytest.mark.unit
def test_pdf_render_produces_valid_pdf(context_with_items) -> None:
    raw = render_pdf(context_with_items)
    assert isinstance(raw, bytes)
    # PDF magic.
    assert raw.startswith(b"%PDF-")
    # Reasonable size for a single-page report with a 3-row table.
    assert len(raw) > 1500


@pytest.mark.unit
def test_context_estimated_savings_sums_cut_items(context_with_items) -> None:
    # Only the Lacework row is "cut" with cost=120k.
    assert context_with_items.estimated_savings == 120_000
    assert context_with_items.savings_cost_known is True
    assert context_with_items.total_cost == 950_000


@pytest.mark.unit
def test_context_marks_savings_unknown_when_cut_has_no_cost() -> None:
    cap_list = CapabilityList(
        id=uuid.uuid4(), service_id=uuid.uuid4(), version=1, status=CapabilityListStatus.DRAFT
    )
    items = [
        _item(annual_cost_usd=None, disposition=CapabilityDisposition.CUT),
        _item(annual_cost_usd=50_000, disposition=CapabilityDisposition.CUT),
    ]
    ctx = build_context(client_legal_name=None, service_title="X", cap_list=cap_list, items=items)
    assert ctx.estimated_savings == 50_000
    assert ctx.savings_cost_known is False
    assert ctx.client_legal_name == "Client"  # fallback when None
