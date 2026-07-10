"""CSF action plan / POA&M routes + export (Playbook Step 10, FIX H-8)."""

from __future__ import annotations

import io
import os
import uuid as _uuid
from collections.abc import Iterator
from pathlib import Path

import pytest
from alembic import command
from alembic.config import Config
from app.csf.catalog import SUBCATEGORIES
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from tests.conftest import register_admin_resp


@pytest.fixture()
def app_client(tmp_path) -> Iterator[TestClient]:
    url = f"sqlite:///{tmp_path / 'shield-csfaction.db'}"
    os.environ["DATABASE_URL"] = url
    api_root = Path(__file__).resolve().parents[2]
    cfg = Config(str(api_root / "alembic.ini"))
    cfg.set_main_option("script_location", str(api_root / "alembic"))
    cfg.set_main_option("sqlalchemy.url", url)
    command.upgrade(cfg, "head")
    engine = create_engine(url, future=True)
    TestSession = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    from app.db.session import get_db
    from app.main import create_app
    from app.models.client import Client as _Client
    from app.models.client_domain import ClientDomain as _ClientDomain
    from app.routes.artifacts import _storage_dep
    from app.storage.local import LocalFilesystemStorage

    def override_get_db() -> Iterator[Session]:
        db = TestSession()
        try:
            yield db
        finally:
            db.close()

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[_storage_dep] = lambda: LocalFilesystemStorage(tmp_path / "storage")
    _seed = TestSession()
    tenant = _Client(legal_name="Test Tenant")
    _seed.add(tenant)
    _seed.flush()
    _seed.add(_ClientDomain(client_id=tenant.id, domain="example.com"))
    _seed.commit()
    cid = str(tenant.id)
    with TestClient(app, headers={"X-Client-Id": cid}) as c:
        yield c, cid


def _admin_headers(c: TestClient) -> dict:
    r = register_admin_resp(c, "admin@example.com")
    return {"Authorization": f"Bearer {r.json()['tokens']['access_token']}"}


def _service_and_assessment(c: TestClient, h: dict) -> tuple[str, dict]:
    svc_id = c.post("/csf/services", headers=h, json={"kind": "nist_csf", "title": "CSF"}).json()[
        "id"
    ]
    assessment = c.post(f"/csf/services/{svc_id}/assessments", headers=h).json()
    return svc_id, assessment


def _score_all(c: TestClient, h: dict, svc_id: str, tier: str) -> int:
    rows = c.get(f"/csf/services/{svc_id}/profile/{tier}", headers=h).json()["rows"]
    for row in rows:
        r = c.patch(
            f"/csf/dimension-scores/{row['id']}",
            headers=h,
            json={
                "governance": 2,
                "policy": 2,
                "implementation": 2,
                "has_evidence": True,
                "target_level": 4,
            },
        )
        assert r.status_code == 200, r.text
    return len(rows)


@pytest.mark.unit
def test_action_item_lifecycle_create_list_patch_delete(app_client) -> None:
    """Create from a gap row, list, PATCH status open -> in_progress -> done,
    then DELETE."""
    c, _ = app_client
    h = _admin_headers(c)
    _svc_id, assessment = _service_and_assessment(c, h)
    aid = assessment["id"]
    code = SUBCATEGORIES[0].code

    # Create from a gap row (the workspace passes the gap's subcategory code).
    created = c.post(
        f"/csf/assessments/{aid}/action-items",
        headers=h,
        json={
            "subcategory_code": code,
            "owner": "Dana Ops",
            "due_date": "2026-09-30",
            "milestone": "Draft and ratify the policy.",
        },
    )
    assert created.status_code == 201, created.text
    item = created.json()
    assert item["subcategory_code"] == code
    assert item["owner"] == "Dana Ops"
    assert item["due_date"] == "2026-09-30"
    assert item["status"] == "open"
    item_id = item["id"]

    # List.
    listed = c.get(f"/csf/assessments/{aid}/action-items", headers=h)
    assert listed.status_code == 200, listed.text
    assert [x["id"] for x in listed.json()] == [item_id]

    # PATCH: open -> in_progress -> done.
    for nxt in ("in_progress", "done"):
        patched = c.patch(f"/csf/action-items/{item_id}", headers=h, json={"status": nxt})
        assert patched.status_code == 200, patched.text
        assert patched.json()["status"] == nxt

    # DELETE.
    deleted = c.delete(f"/csf/action-items/{item_id}", headers=h)
    assert deleted.status_code == 204, deleted.text
    assert c.get(f"/csf/assessments/{aid}/action-items", headers=h).json() == []


@pytest.mark.unit
def test_action_item_empty_patch_400(app_client) -> None:
    c, _ = app_client
    h = _admin_headers(c)
    _svc_id, assessment = _service_and_assessment(c, h)
    created = c.post(
        f"/csf/assessments/{assessment['id']}/action-items",
        headers=h,
        json={"subcategory_code": SUBCATEGORIES[0].code},
    )
    item_id = created.json()["id"]
    assert c.patch(f"/csf/action-items/{item_id}", headers=h, json={}).status_code == 400


@pytest.mark.unit
def test_action_item_unknown_subcategory_422(app_client) -> None:
    c, _ = app_client
    h = _admin_headers(c)
    _svc_id, assessment = _service_and_assessment(c, h)
    r = c.post(
        f"/csf/assessments/{assessment['id']}/action-items",
        headers=h,
        json={"subcategory_code": "ZZ.NOPE-99"},
    )
    assert r.status_code == 422, r.text


@pytest.mark.unit
def test_action_item_cross_tenant_is_404_not_403(app_client) -> None:
    """Another tenant's assessment id must 404 (no existence oracle), never 403
    or 500 — on both the create and the list route."""
    c, _cid = app_client
    h = _admin_headers(c)

    # Stand up a second tenant + an assessment owned by it.
    engine = create_engine(os.environ["DATABASE_URL"], future=True)
    from app.models.client import Client as _Client

    with Session(engine, future=True) as s:
        other = _Client(legal_name="Other Tenant")
        s.add(other)
        s.commit()
        other_id = str(other.id)
    engine.dispose()

    other_h = {**h, "X-Client-Id": other_id}
    other_svc = c.post(
        "/csf/services", headers=other_h, json={"kind": "nist_csf", "title": "Other CSF"}
    ).json()["id"]
    other_assessment = c.post(f"/csf/services/{other_svc}/assessments", headers=other_h).json()[
        "id"
    ]

    # As tenant1 (the baked-in X-Client-Id header), tenant2's assessment 404s.
    assert c.get(f"/csf/assessments/{other_assessment}/action-items", headers=h).status_code == 404
    assert (
        c.post(
            f"/csf/assessments/{other_assessment}/action-items",
            headers=h,
            json={"subcategory_code": SUBCATEGORIES[0].code},
        ).status_code
        == 404
    )
    # A wholly unknown assessment id 404s too (not 500).
    assert c.get(f"/csf/assessments/{_uuid.uuid4()}/action-items", headers=h).status_code == 404


@pytest.mark.unit
def test_patch_delete_cross_tenant_item_is_404(app_client) -> None:
    """A tenant cannot PATCH/DELETE another tenant's action item."""
    c, _cid = app_client
    h = _admin_headers(c)

    engine = create_engine(os.environ["DATABASE_URL"], future=True)
    from app.models.client import Client as _Client

    with Session(engine, future=True) as s:
        other = _Client(legal_name="Other Tenant")
        s.add(other)
        s.commit()
        other_id = str(other.id)
    engine.dispose()

    other_h = {**h, "X-Client-Id": other_id}
    other_svc = c.post(
        "/csf/services", headers=other_h, json={"kind": "nist_csf", "title": "Other CSF"}
    ).json()["id"]
    other_assessment = c.post(f"/csf/services/{other_svc}/assessments", headers=other_h).json()[
        "id"
    ]
    other_item = c.post(
        f"/csf/assessments/{other_assessment}/action-items",
        headers=other_h,
        json={"subcategory_code": SUBCATEGORIES[0].code},
    ).json()["id"]

    # tenant1 can't touch tenant2's item.
    assert (
        c.patch(f"/csf/action-items/{other_item}", headers=h, json={"status": "done"}).status_code
        == 404
    )
    assert c.delete(f"/csf/action-items/{other_item}", headers=h).status_code == 404


@pytest.mark.unit
def test_creating_action_item_does_not_set_scored_at_and_b3_gate_holds(app_client) -> None:
    """B-3 interaction: creating an action item must not score the assessment.
    An unscored (seeded) assessment with action items still 409s on export."""
    c, _ = app_client
    h = _admin_headers(c)
    svc_id, assessment = _service_and_assessment(c, h)
    c.post(f"/csf/services/{svc_id}/profiles/seed", headers=h, json={"tiers": ["high"]})

    # Add an action item on the still-unscored assessment.
    r = c.post(
        f"/csf/assessments/{assessment['id']}/action-items",
        headers=h,
        json={"subcategory_code": SUBCATEGORIES[0].code, "owner": "Dana"},
    )
    assert r.status_code == 201, r.text

    # Export is STILL blocked: the action item touched no dimension row, so
    # scored_at stayed NULL for every in-scope row.
    blocked = c.post(f"/csf/services/{svc_id}/playbook/export", headers=h)
    assert blocked.status_code == 409, blocked.text
    n = len(SUBCATEGORIES)
    assert str(n) in blocked.json()["error"]["message"], blocked.json()


@pytest.mark.unit
def test_action_plan_sheet_mirrors_created_items(app_client) -> None:
    """The exported XLSX has an 'Action Plan' sheet whose rows MIRROR the created
    items (subcategory, owner, due date, milestone, status) — asserted on real
    openpyxl cell values, not just HTTP 200."""
    c, cid = app_client
    h = _admin_headers(c)
    svc_id, assessment = _service_and_assessment(c, h)
    c.post(f"/csf/services/{svc_id}/profiles/seed", headers=h, json={"tiers": ["high"]})
    n = _score_all(c, h, svc_id, "high")
    assert n == len(SUBCATEGORIES)
    c.post(f"/csf/assessments/{assessment['id']}/approve", headers=h)

    aid = assessment["id"]
    code_a = SUBCATEGORIES[0].code
    code_b = SUBCATEGORIES[1].code
    # Two items with distinct subcategory codes so the export ordering
    # (subcategory_code, created_at) is deterministic.
    c.post(
        f"/csf/assessments/{aid}/action-items",
        headers=h,
        json={
            "subcategory_code": code_a,
            "owner": "Dana Ops",
            "due_date": "2026-09-30",
            "milestone": "Ratify the policy.",
            "status": "in_progress",
        },
    )
    c.post(
        f"/csf/assessments/{aid}/action-items",
        headers=h,
        json={"subcategory_code": code_b, "owner": "Sam Sec", "due_date": "2026-12-01"},
    )

    ex = c.post(f"/csf/services/{svc_id}/playbook/export", headers=h)
    assert ex.status_code == 200, ex.text
    xlsx = next(a for a in ex.json()["artifacts"] if a["kind"] == "xlsx")

    dh = {**h, "X-Client-Id": cid}
    dl = c.get(f"/artifacts/{xlsx['artifact_id']}/download", headers=dh)
    assert dl.status_code == 200, dl.text
    wb = load_workbook(io.BytesIO(dl.content))
    assert "Action Plan" in wb.sheetnames, wb.sheetnames
    ws = wb["Action Plan"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("Subcategory", "Owner", "Due date", "Milestone", "Status")
    # Ordered by subcategory_code: code_a (index 0) then code_b (index 1).
    expected = sorted([code_a, code_b])
    body = rows[1:]
    assert [r[0] for r in body] == expected
    by_code = {r[0]: r for r in body}
    assert by_code[code_a] == (
        code_a,
        "Dana Ops",
        "2026-09-30",
        "Ratify the policy.",
        "In progress",
    )
    # openpyxl reads an empty-string cell back as None; the milestone was blank.
    assert by_code[code_b] == (code_b, "Sam Sec", "2026-12-01", None, "Open")
