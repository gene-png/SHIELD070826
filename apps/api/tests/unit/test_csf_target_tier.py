"""FIX B-2: CSF finalize must honor the client's chosen target tier.

The old finalize passed no target to analyze_gaps, so it always measured against
DEFAULT_TARGET_TIER (3). A client targeting T4 got an exported Gap Plan that
disagreed with the workspace gap list (which does honor the target). This test
builds an assessment whose originating ServiceRequest targets T4, scores every
subcategory at T3, and asserts the finalized XLSX Gap Plan row count equals the
workspace (gap-analysis endpoint) gap count at T4, is non-zero, and that the
deliverable summary states the resolved target.
"""

from __future__ import annotations

import io
import os
from collections.abc import Iterator
from pathlib import Path

import pytest
from alembic import command
from alembic.config import Config
from app.storage.local import LocalFilesystemStorage
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from tests.conftest import register_admin


@pytest.fixture()
def app_client(tmp_path) -> Iterator[tuple[TestClient, sessionmaker, str]]:
    url = f"sqlite:///{tmp_path / 'shield-csftgt.db'}"
    os.environ["DATABASE_URL"] = url
    api_root = Path(__file__).resolve().parents[2]
    cfg = Config(str(api_root / "alembic.ini"))
    cfg.set_main_option("script_location", str(api_root / "alembic"))
    cfg.set_main_option("sqlalchemy.url", url)
    command.upgrade(cfg, "head")

    engine = create_engine(url, future=True)
    TestSession = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    storage = LocalFilesystemStorage(tmp_path / "storage")

    from app.db.session import get_db
    from app.main import create_app
    from app.models.client import Client as _Client
    from app.models.client_domain import ClientDomain as _ClientDomain
    from app.routes.artifacts import _storage_dep

    def override_get_db() -> Iterator[Session]:
        db = TestSession()
        try:
            yield db
        finally:
            db.close()

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[_storage_dep] = lambda: storage
    _seed = TestSession()
    tenant = _Client(legal_name="Test Tenant")
    _seed.add(tenant)
    _seed.flush()
    _seed.add(_ClientDomain(client_id=tenant.id, domain="example.com"))
    _seed.commit()
    cid = str(tenant.id)
    _seed.close()

    with TestClient(app, headers={"X-Client-Id": cid}) as c:
        yield c, TestSession, cid


@pytest.mark.unit
def test_finalize_honors_client_target_tier(app_client) -> None:
    c, TestSession, cid = app_client
    admin = register_admin(c, "admin@example.com")
    bearer = admin["tokens"]["access_token"]
    admin_user_id = admin["user"]["id"]
    h = {"Authorization": f"Bearer {bearer}"}

    # A ServiceRequest that pins the client's CSF target tier to 4.
    import uuid as _uuid

    from app.models.service_request import ServiceRequest, ServiceType

    seed = TestSession()
    sr = ServiceRequest(
        service_type=ServiceType.NIST_CSF,
        client_id=_uuid.UUID(cid),
        requested_by=_uuid.UUID(admin_user_id),
        csf_target_tier=4,
        csf_profile="HIGH",
    )
    seed.add(sr)
    seed.commit()
    sr_id = str(sr.id)
    seed.close()

    svc_id = c.post(
        "/csf/services",
        headers=h,
        json={"kind": "nist_csf", "title": "Atlas - CSF", "source_request_id": sr_id},
    ).json()["id"]
    assess = c.post(f"/csf/services/{svc_id}/assessments", headers=h).json()
    # Score every subcategory at T3 (below the T4 target -> every row is a gap).
    for ans in assess["answers"]:
        c.patch(f"/csf/answers/{ans['id']}", headers=h, json={"maturity_tier": 3})
    c.post(f"/csf/assessments/{assess['id']}/approve", headers=h)

    # Workspace/dashboard gap list at the client's real target (T4).
    ga = c.get(f"/csf/services/{svc_id}/gap-analysis?target_tier=4", headers=h).json()
    assert ga["target_tier"] == 4
    assert ga["total_gap_count"] > 0
    workspace_gap_rows = len(ga["gaps"])
    assert workspace_gap_rows > 0

    fin = c.post(f"/csf/services/{svc_id}/deliverables/finalize", headers=h)
    assert fin.status_code == 201, fin.text
    body = fin.json()
    # The deliverable states its own assumption: the resolved target tier.
    assert "T4" in body["summary"], body["summary"]

    dl = c.get(f"/artifacts/{body['xlsx_artifact_id']}/download", headers=h)
    assert dl.status_code == 200
    wb = load_workbook(io.BytesIO(dl.content))
    gap_ws = wb["Gap Plan"]
    xlsx_gap_rows = gap_ws.max_row - 1  # minus the header row

    # The exported Gap Plan matches the workspace gap list (same target tier) and
    # is non-zero. With the pre-fix code (target hardcoded to T3, all rows at T3)
    # there are ZERO gaps, so the exported sheet carries a single "No gaps"
    # placeholder row instead — a mismatch this assertion catches.
    assert xlsx_gap_rows == workspace_gap_rows
    assert xlsx_gap_rows > 0
