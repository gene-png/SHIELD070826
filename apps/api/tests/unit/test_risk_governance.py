"""Risk Register governance (FIX F-3): review/edit/lock/delete/approve + gate.

The Risk Register was the one flow with too few steps: AI-generated entries flowed
straight into the client-facing export with no edit/delete/lock, and the gate only
checked that source assessments EXISTED (they auto-create at intake) rather than
that they were APPROVED. These tests pin the added governance:

  * PATCH an entry; the tier stays code-derived, a client-supplied tier is ignored.
  * Lock an entry; regenerate preserves it verbatim and redrafts only unlocked ones.
  * DELETE (soft) drops an entry from the register and every export.
  * The gate refuses on un-approved sources and opens on approved ones.
  * CSF harvest honours the client's chosen target tier (not a fixed below-3).
  * Export is refused before approve; after approve it renders, and an edited
    entry appears in the real XLSX.
  * Risk export filenames follow the §15.5 deliverable_filename convention.
"""

from __future__ import annotations

import io
import os
import uuid
from collections.abc import Iterator
from datetime import date
from pathlib import Path

import pytest
from alembic import command
from alembic.config import Config
from app.ai.llm import FixtureProvider, LLMClient, LLMResponse
from app.tech_debt.filename import deliverable_filename
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from tests.conftest import register_admin_resp


@pytest.fixture()
def app_client(tmp_path) -> Iterator[tuple[TestClient, FixtureProvider, sessionmaker]]:
    url = f"sqlite:///{tmp_path / 'shield-risk-gov.db'}"
    os.environ["DATABASE_URL"] = url
    api_root = Path(__file__).resolve().parents[2]
    cfg = Config(str(api_root / "alembic.ini"))
    cfg.set_main_option("script_location", str(api_root / "alembic"))
    cfg.set_main_option("sqlalchemy.url", url)
    command.upgrade(cfg, "head")
    engine = create_engine(url, future=True)
    TestSession = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    from app.db.session import get_db
    from app.main import create_app
    from app.routes.artifacts import _storage_dep
    from app.routes.risk import _llm_dep
    from app.storage.local import LocalFilesystemStorage

    def override_get_db() -> Iterator[Session]:
        db = TestSession()
        try:
            yield db
        finally:
            db.close()

    provider = FixtureProvider()
    storage = LocalFilesystemStorage(tmp_path / "storage")
    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[_llm_dep] = lambda: LLMClient(provider)
    app.dependency_overrides[_storage_dep] = lambda: storage
    with TestClient(app) as c:
        yield c, provider, TestSession


def _admin(c: TestClient) -> tuple[str, str, str]:
    admin = register_admin_resp(c, "admin@kentro.example")
    body = admin.json()
    bearer = body["tokens"]["access_token"]
    uid = body["user"]["id"]
    cid = c.post(
        "/admin/clients",
        headers={"Authorization": f"Bearer {bearer}"},
        json={"legal_name": "Acme"},
    ).json()["id"]
    return bearer, cid, uid


def _seed_sources(c: TestClient, bearer: str, cid: str, *, approve: bool = True) -> tuple[str, str]:
    """Create an ATT&CK + a ZT assessment; optionally approve both."""
    h = {"Authorization": f"Bearer {bearer}", "X-Client-Id": cid}
    asvc = c.post("/attack/services", headers=h, json={"kind": "attack_coverage", "title": "A"})
    a = c.post(f"/attack/services/{asvc.json()['id']}/assessments", headers=h)
    cov = a.json()["coverage"][0]
    technique = cov["technique_code"]
    c.patch(f"/attack/coverage/{cov['id']}", headers=h, json={"status": "gap"})

    zsvc = c.post("/zt/services", headers=h, json={"kind": "zero_trust_cisa", "title": "Z"})
    za = c.post(f"/zt/services/{zsvc.json()['id']}/assessments", headers=h)
    zans = za.json()["answers"][0]
    capability = zans["capability_code"]
    c.patch(f"/zt/answers/{zans['id']}", headers=h, json={"maturity_stage": 1})

    if approve:
        c.post(f"/attack/assessments/{a.json()['id']}/approve", headers=h)
        c.post(f"/zt/assessments/{za.json()['id']}/approve", headers=h)
    return technique, capability


# ---------------------------------------------------------------------------
# Gate: existence is not enough; approval is
# ---------------------------------------------------------------------------


@pytest.mark.unit
def test_gate_refuses_when_sources_exist_but_unapproved(app_client) -> None:
    c, _provider, _S = app_client
    bearer, cid, _uid = _admin(c)
    bh = {"Authorization": f"Bearer {bearer}"}
    _seed_sources(c, bearer, cid, approve=False)  # started, NOT approved

    g = c.get(f"/risk/clients/{cid}/gate", headers=bh).json()
    assert g["has_attack"] is True and g["has_zt"] is True  # they exist...
    assert g["attack_approved"] is False and g["zt_approved"] is False
    assert g["unlocked"] is False  # ...but the gate stays shut
    # Generate refuses too.
    r = c.post(f"/risk/clients/{cid}/register/generate", headers=bh)
    assert r.status_code == 409, r.text


@pytest.mark.unit
def test_gate_opens_on_attack_plus_one_approved(app_client) -> None:
    c, _provider, _S = app_client
    bearer, cid, _uid = _admin(c)
    bh = {"Authorization": f"Bearer {bearer}"}
    _seed_sources(c, bearer, cid, approve=True)

    g = c.get(f"/risk/clients/{cid}/gate", headers=bh).json()
    assert g["attack_approved"] is True
    assert g["zt_approved"] is True
    assert g["unlocked"] is True
    # Sources are labelled with their status for the dashboard.
    kinds = {s["kind"]: s for s in g["sources"]}
    assert kinds["attack"]["approved"] is True
    assert kinds["attack"]["status"] == "approved"


# ---------------------------------------------------------------------------
# PATCH: tier stays code-derived, client tier ignored
# ---------------------------------------------------------------------------


def _generate_one(c: TestClient, provider: FixtureProvider, bearer: str, cid: str) -> dict:
    bh = {"Authorization": f"Bearer {bearer}"}
    provider.register_static(
        "risk_synthesize",
        LLMResponse(
            '{"entries": [{"title": "Initial risk", "description": "d",'
            ' "axis": "detection", "likelihood": "low", "impact": "minor",'
            ' "recommended_action": "mitigate"}]}'
        ),
    )
    r = c.post(f"/risk/clients/{cid}/register/generate", headers=bh)
    assert r.status_code == 201, r.text
    return r.json()


@pytest.mark.unit
def test_patch_rederives_tier_and_ignores_client_tier(app_client) -> None:
    c, provider, _S = app_client
    bearer, cid, _uid = _admin(c)
    bh = {"Authorization": f"Bearer {bearer}"}
    _seed_sources(c, bearer, cid)
    body = _generate_one(c, provider, bearer, cid)
    e = body["entries"][0]
    # low x minor -> score (1+1)*(1+1)=4 -> Low.
    assert e["tier"] == "low"

    # Edit title/description/likelihood/impact AND smuggle a bogus tier.
    r = c.patch(
        f"/risk/entries/{e['id']}",
        headers=bh,
        json={
            "title": "Edited weakness",
            "description": "new desc",
            "likelihood": "high",
            "impact": "catastrophic",
            "tier": "negligible",  # must be IGNORED
        },
    )
    assert r.status_code == 200, r.text
    out = r.json()
    assert out["title"] == "Edited weakness"
    assert out["description"] == "new desc"
    assert out["likelihood"] == "high"
    assert out["impact"] == "catastrophic"
    # high x catastrophic -> Critical (code-derived), NOT the client's "negligible".
    assert out["tier"] == "critical"


# ---------------------------------------------------------------------------
# Lock + regenerate: locked survives verbatim, unlocked redrafted
# ---------------------------------------------------------------------------


@pytest.mark.unit
def test_regenerate_preserves_locked_entries(app_client) -> None:
    c, provider, _S = app_client
    bearer, cid, _uid = _admin(c)
    bh = {"Authorization": f"Bearer {bearer}"}
    _seed_sources(c, bearer, cid)

    provider.register_static(
        "risk_synthesize",
        LLMResponse(
            '{"entries": [{"title": "Draft one", "axis": "detection",'
            ' "likelihood": "high", "impact": "catastrophic",'
            ' "recommended_action": "remediate"}]}'
        ),
    )
    v1 = c.post(f"/risk/clients/{cid}/register/generate", headers=bh).json()
    e = v1["entries"][0]
    # Lock it and rename so we can spot it verbatim across the regenerate.
    c.patch(
        f"/risk/entries/{e['id']}",
        headers=bh,
        json={
            "title": "Locked risk",
            "likelihood": "high",
            "impact": "catastrophic",
            "locked": True,
        },
    )

    # A different AI draft for the next generation.
    provider.register_static(
        "risk_synthesize",
        LLMResponse(
            '{"entries": [{"title": "Draft two", "axis": "prevention",'
            ' "likelihood": "low", "impact": "minor", "recommended_action": "accept"}]}'
        ),
    )
    v2 = c.post(f"/risk/clients/{cid}/register/generate", headers=bh).json()
    assert v2["version"] == 2
    titles = {x["title"]: x for x in v2["entries"]}
    # Locked entry carried forward VERBATIM (tier still code-derived critical).
    assert "Locked risk" in titles
    locked_entry = titles["Locked risk"]
    assert locked_entry["locked"] is True
    assert locked_entry["tier"] == "critical"
    # The unlocked v1 draft was redrafted away; the new AI draft is present.
    assert "Draft one" not in titles
    assert "Draft two" in titles


# ---------------------------------------------------------------------------
# DELETE (soft): disappears from register + export
# ---------------------------------------------------------------------------


@pytest.mark.unit
def test_soft_delete_removes_entry_from_register_and_export(app_client) -> None:
    c, provider, _S = app_client
    bearer, cid, _uid = _admin(c)
    bh = {"Authorization": f"Bearer {bearer}"}
    _seed_sources(c, bearer, cid)
    provider.register_static(
        "risk_synthesize",
        LLMResponse(
            '{"entries": ['
            '{"title": "Keep me", "axis": "detection", "likelihood": "high",'
            ' "impact": "catastrophic", "recommended_action": "remediate"},'
            '{"title": "Delete me", "axis": "response", "likelihood": "low",'
            ' "impact": "minor", "recommended_action": "accept"}]}'
        ),
    )
    v1 = c.post(f"/risk/clients/{cid}/register/generate", headers=bh).json()
    by_title = {x["title"]: x for x in v1["entries"]}
    assert set(by_title) == {"Keep me", "Delete me"}

    d = c.delete(f"/risk/entries/{by_title['Delete me']['id']}", headers=bh)
    assert d.status_code == 204, d.text

    latest = c.get(f"/risk/clients/{cid}/register/latest", headers=bh).json()
    remaining = {x["title"] for x in latest["entries"]}
    assert remaining == {"Keep me"}

    # And it is absent from the exported XLSX.
    c.post(f"/risk/clients/{cid}/register/approve", headers=bh)
    exp = c.post(f"/risk/clients/{cid}/register/export", headers=bh).json()
    dh = {**bh, "X-Client-Id": cid}
    xlsx = c.get(f"/artifacts/{exp['xlsx_artifact_id']}/download", headers=dh)
    wb = load_workbook(io.BytesIO(xlsx.content))
    cells = {cell.value for row in wb.active.iter_rows() for cell in row}
    assert "Keep me" in cells
    assert "Delete me" not in cells


# ---------------------------------------------------------------------------
# Export before approve refused; after approve renders + carries the edit
# ---------------------------------------------------------------------------


@pytest.mark.unit
def test_export_requires_approve_and_carries_edit(app_client) -> None:
    c, provider, _S = app_client
    bearer, cid, _uid = _admin(c)
    bh = {"Authorization": f"Bearer {bearer}"}
    _seed_sources(c, bearer, cid)
    body = _generate_one(c, provider, bearer, cid)
    e = body["entries"][0]
    c.patch(
        f"/risk/entries/{e['id']}",
        headers=bh,
        json={
            "title": "Consultant-edited exposure",
            "likelihood": "high",
            "impact": "catastrophic",
        },
    )

    # Export BEFORE approve -> refused.
    early = c.post(f"/risk/clients/{cid}/register/export", headers=bh)
    assert early.status_code == 409, early.text

    # Approve, then export succeeds.
    ap = c.post(f"/risk/clients/{cid}/register/approve", headers=bh)
    assert ap.status_code == 200 and ap.json()["approved_at"] is not None
    exp = c.post(f"/risk/clients/{cid}/register/export", headers=bh)
    assert exp.status_code == 200, exp.text
    out = exp.json()

    # Filenames follow the §15.5 deliverable_filename convention.
    expected_xlsx = deliverable_filename(
        company="Acme", service_slug="Risk_Register", extension="xlsx", day=date.today(), version=1
    )
    assert out["xlsx_filename"] == expected_xlsx
    assert out["pdf_filename"] == expected_xlsx[:-4] + "pdf"
    assert out["docx_filename"] == expected_xlsx[:-4] + "docx"

    # Open the real XLSX; the edited title is present (not the original).
    dh = {**bh, "X-Client-Id": cid}
    xlsx = c.get(f"/artifacts/{out['xlsx_artifact_id']}/download", headers=dh)
    assert xlsx.status_code == 200 and xlsx.content[:2] == b"PK"
    wb = load_workbook(io.BytesIO(xlsx.content))
    cells = {cell.value for row in wb.active.iter_rows() for cell in row}
    assert "Consultant-edited exposure" in cells
    assert "Initial risk" not in cells


@pytest.mark.unit
def test_approved_version_is_locked_against_edits(app_client) -> None:
    c, provider, _S = app_client
    bearer, cid, _uid = _admin(c)
    bh = {"Authorization": f"Bearer {bearer}"}
    _seed_sources(c, bearer, cid)
    body = _generate_one(c, provider, bearer, cid)
    e = body["entries"][0]
    c.post(f"/risk/clients/{cid}/register/approve", headers=bh)
    # Editing an approved version is refused.
    r = c.patch(f"/risk/entries/{e['id']}", headers=bh, json={"title": "too late"})
    assert r.status_code == 409, r.text


# ---------------------------------------------------------------------------
# CSF harvest honours the client's target tier (unit on _gather_findings)
# ---------------------------------------------------------------------------


@pytest.mark.unit
def test_csf_harvest_honours_client_target_tier(app_client) -> None:
    c, _provider, TestSession = app_client
    bearer, cid, uid = _admin(c)

    from app.models.csf_assessment import CsfAnswer, CsfAssessment, CsfAssessmentStatus
    from app.models.service import Service, ServiceKind, ServiceStatus
    from app.models.service_request import ServiceRequest, ServiceType
    from app.routes.risk import _gather_findings

    def _build_csf(target_tier: int, answer_tier: int) -> str:
        """Create a CSF service/assessment for a fresh code and return that code."""
        code = f"GV.OC-{uuid.uuid4().hex[:2]}"
        with TestSession() as s:
            sr = ServiceRequest(
                service_type=ServiceType.NIST_CSF,
                client_id=uuid.UUID(cid),
                requested_by=uuid.UUID(uid),
                csf_target_tier=target_tier,
                csf_profile="HIGH",
            )
            s.add(sr)
            s.flush()
            svc = Service(
                kind=ServiceKind.NIST_CSF,
                status=ServiceStatus.IN_PROGRESS,
                title="CSF",
                client_id=uuid.UUID(cid),
                source_request_id=sr.id,
                opened_by=uuid.UUID(uid),
            )
            s.add(svc)
            s.flush()
            a = CsfAssessment(
                service_id=svc.id,
                client_id=uuid.UUID(cid),
                version=1,
                status=CsfAssessmentStatus.APPROVED,
            )
            s.add(a)
            s.flush()
            s.add(
                CsfAnswer(
                    assessment_id=a.id,
                    client_id=uuid.UUID(cid),
                    subcategory_code=code,
                    maturity_tier=answer_tier,
                )
            )
            s.commit()
        return code

    # Target 4, answer at tier 3: 3 < 4 -> the subcategory IS harvested. The old
    # fixed below-tier-3 threshold (3 < 3 is False) would have MISSED this.
    code = _build_csf(target_tier=4, answer_tier=3)
    with TestSession() as s:
        findings, _tech, _ctrl = _gather_findings(s, uuid.UUID(cid))
    csf_ids = {f["source_id"] for f in findings if f["kind"] == "csf"}
    assert code in csf_ids


@pytest.mark.unit
def test_csf_harvest_default_threshold_is_three(app_client) -> None:
    """No target pinned -> falls back to below-tier-3 (a tier-3 answer is clean)."""
    c, _provider, TestSession = app_client
    bearer, cid, uid = _admin(c)

    from app.models.csf_assessment import CsfAnswer, CsfAssessment, CsfAssessmentStatus
    from app.routes.risk import _gather_findings

    code_low = f"GV.OC-{uuid.uuid4().hex[:2]}"
    code_ok = f"GV.OC-{uuid.uuid4().hex[:2]}"
    with TestSession() as s:
        # No Service/ServiceRequest -> target resolves to the default (3).
        a = CsfAssessment(
            service_id=uuid.uuid4(),  # dangling on purpose: _client_target_tier -> None -> 3
            client_id=uuid.UUID(cid),
            version=1,
            status=CsfAssessmentStatus.APPROVED,
        )
        s.add(a)
        s.flush()
        s.add(
            CsfAnswer(
                assessment_id=a.id,
                client_id=uuid.UUID(cid),
                subcategory_code=code_low,
                maturity_tier=2,
            )
        )
        s.add(
            CsfAnswer(
                assessment_id=a.id,
                client_id=uuid.UUID(cid),
                subcategory_code=code_ok,
                maturity_tier=3,
            )
        )
        s.commit()
    with TestSession() as s:
        findings, _t, _ctrl = _gather_findings(s, uuid.UUID(cid))
    csf_ids = {f["source_id"] for f in findings if f["kind"] == "csf"}
    assert code_low in csf_ids  # tier 2 < 3 -> harvested
    assert code_ok not in csf_ids  # tier 3 is at target -> clean
