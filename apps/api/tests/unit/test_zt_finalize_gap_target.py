"""Regression test for defect B-1.

The ZT dashboard gap endpoint resolves the client's real target stage (the
engagement-level goal captured at intake, plus any per-capability targets). The
finalize path used to ignore all of that and fall back to the hardcoded default
target (stage 3), so the signed client deliverable could report
"0 gap(s) at target S3" with a "No gaps at target stage" Gap Plan sheet while
the consultant's screen showed 37 gaps at target S4.

This test pins the two together: a client whose intake target is stage 4, with
every capability scored at stage 3, must produce a finalized artifact whose gap
count EQUALS the dashboard's ``total_gap_count`` (and is not zero), whose
summary states the resolved target (S4, not the default S3), and whose actual
XLSX Gap Plan bytes are not the empty placeholder.
"""

from __future__ import annotations

import io
import os
import re
from collections.abc import Iterator
from pathlib import Path

import pytest
from alembic import command
from alembic.config import Config
from app.models.client import Client
from app.models.client_domain import ClientDomain
from app.storage.local import LocalFilesystemStorage
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from tests.conftest import register_admin


@pytest.fixture()
def app_client(tmp_path) -> Iterator[TestClient]:
    db_path = tmp_path / "shield-zt-b1.db"
    url = f"sqlite:///{db_path}"
    os.environ["DATABASE_URL"] = url
    api_root = Path(__file__).resolve().parents[2]
    cfg = Config(str(api_root / "alembic.ini"))
    cfg.set_main_option("script_location", str(api_root / "alembic"))
    cfg.set_main_option("sqlalchemy.url", url)
    command.upgrade(cfg, "head")

    engine = create_engine(url, future=True)
    TestSession = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    storage = LocalFilesystemStorage(tmp_path / "storage")

    from app.db.session import get_db
    from app.main import create_app
    from app.routes.artifacts import _storage_dep

    def override_get_db() -> Iterator[Session]:
        db = TestSession()
        try:
            yield db
        finally:
            db.close()

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[_storage_dep] = lambda: storage

    # Work Order B1: seed a "(pending intake)" client + approved domain so the
    # client-role registrant auto-joins it and can run intake.
    seed = TestSession()
    tenant = Client(legal_name="(pending intake)")
    seed.add(tenant)
    seed.flush()
    seed.add(ClientDomain(client_id=tenant.id, domain="example.com"))
    seed.commit()
    seed.close()

    with TestClient(app) as c:
        yield c


def _register_client(c: TestClient, email: str) -> dict:
    r = c.post(
        "/auth/register",
        json={
            "email": email,
            "password": "correct horse battery staple!",
            "display_name": email.split("@")[0],
        },
    )
    assert r.status_code == 201, r.text
    return r.json()


@pytest.mark.unit
def test_finalized_gap_count_matches_dashboard_not_default_target(app_client) -> None:
    c = app_client
    admin = register_admin(c, "admin@example.com")
    bearer_admin = admin["tokens"]["access_token"]

    # Client runs intake declaring a ZT target of stage 4 (the engagement-level
    # goal the whole deliverable is measured against).
    client = _register_client(c, "client@example.com")
    bearer_client = client["tokens"]["access_token"]
    client_cid = client["user"]["client_id"]
    intake = c.post(
        "/intake/submit",
        headers={"Authorization": f"Bearer {bearer_client}"},
        json={
            "client": {"legal_name": "Atlas Defense Solutions"},
            "service_requests": [
                {"service_type": "zero_trust_cisa", "zt_target_stage": 4},
            ],
        },
    )
    assert intake.status_code == 200, intake.text
    sr = next(
        s for s in intake.json()["service_requests"] if s["service_type"] == "zero_trust_cisa"
    )
    svc_id = sr["fulfilled_service_id"]
    assert svc_id is not None, "intake did not auto-provision the ZT service"

    # Admin operates inside the client's tenant.
    h = {"Authorization": f"Bearer {bearer_admin}", "X-Client-Id": client_cid}

    assessment = c.post(f"/zt/services/{svc_id}/assessments", headers=h).json()
    # Every capability at stage 3 -> a 1-stage gap on all 37 against target S4.
    for ans in assessment["answers"]:
        r = c.patch(
            f"/zt/answers/{ans['id']}",
            headers=h,
            json={"maturity_stage": 3},
        )
        assert r.status_code == 200, r.text

    # 1. Dashboard gap count (what the consultant reviews on screen). No explicit
    #    target override -> it must resolve the client's intake target (S4).
    dash = c.get(f"/zt/services/{svc_id}/gap-analysis", headers=h)
    assert dash.status_code == 200, dash.text
    dash_body = dash.json()
    dash_count = dash_body["total_gap_count"]
    assert dash_body["target_stage"] == 4
    assert dash_count == 37  # all 37 capabilities sit a stage below target S4
    assert dash_count != 0

    # 2. Approve + finalize -> the signed client deliverable.
    ap = c.post(f"/zt/assessments/{assessment['id']}/approve", headers=h)
    assert ap.status_code == 200, ap.text
    fin = c.post(f"/zt/services/{svc_id}/deliverables/finalize", headers=h)
    assert fin.status_code == 201, fin.text
    deliv = fin.json()

    # The finalized artifact's gap count EQUALS the dashboard's and is NOT zero,
    # and the summary states the resolved target (S4), not the default (S3).
    m = re.search(r"(\d+) gap\(s\) at target S(\d+)", deliv["summary"])
    assert m is not None, deliv["summary"]
    finalized_count = int(m.group(1))
    finalized_target = int(m.group(2))
    assert finalized_count == dash_count
    assert finalized_count != 0
    assert finalized_target == 4

    # 3. Strongest assertion: the actual XLSX bytes the client receives.
    xlsx = c.get(f"/artifacts/{deliv['xlsx_artifact_id']}/download", headers=h)
    assert xlsx.status_code == 200, xlsx.text
    wb = load_workbook(io.BytesIO(xlsx.content))
    ws = wb["Gap Plan"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # Not the empty-state placeholder ("No gaps at target stage").
    assert not any(r[2] == "No gaps at target stage" for r in rows), rows
    # Every listed gap targets the resolved stage 4, and there is at least one.
    gap_rows = [r for r in rows if r[0] not in (None, "—")]
    assert gap_rows, rows
    assert all(r[4] == 4 for r in gap_rows), gap_rows
