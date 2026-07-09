"""CSF full-Playbook XLSX export (Work Order D4)."""

from __future__ import annotations

import os
from collections.abc import Iterator
from datetime import UTC
from pathlib import Path

import pytest
from alembic import command
from alembic.config import Config
from app.csf.catalog import SUBCATEGORIES
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from tests.conftest import register_admin_resp


@pytest.fixture()
def app_client(tmp_path) -> Iterator[TestClient]:
    url = f"sqlite:///{tmp_path / 'shield-csfxlsx.db'}"
    os.environ["DATABASE_URL"] = url
    api_root = Path(__file__).resolve().parents[2]
    cfg = Config(str(api_root / "alembic.ini"))
    cfg.set_main_option("script_location", str(api_root / "alembic"))
    cfg.set_main_option("sqlalchemy.url", url)
    command.upgrade(cfg, "head")
    engine = create_engine(url, future=True)
    TestSession = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    from app.db.session import get_db
    from app.main import create_app
    from app.models.client import Client as _Client
    from app.models.client_domain import ClientDomain as _ClientDomain
    from app.routes.artifacts import _storage_dep
    from app.storage.local import LocalFilesystemStorage

    def override_get_db() -> Iterator[Session]:
        db = TestSession()
        try:
            yield db
        finally:
            db.close()

    app = create_app()
    app.dependency_overrides[get_db] = override_get_db
    app.dependency_overrides[_storage_dep] = lambda: LocalFilesystemStorage(tmp_path / "storage")
    _seed = TestSession()
    tenant = _Client(legal_name="Test Tenant")
    _seed.add(tenant)
    _seed.flush()
    _seed.add(_ClientDomain(client_id=tenant.id, domain="example.com"))
    _seed.commit()
    cid = str(tenant.id)
    with TestClient(app, headers={"X-Client-Id": cid}) as c:
        yield c, cid


def _score_all(c: TestClient, h: dict, svc_id: str, tier: str) -> int:
    """Score every seeded row in `tier` at L3 (total 6, evidence present so no
    cap). Returns the number of rows scored."""
    rows = c.get(f"/csf/services/{svc_id}/profile/{tier}", headers=h).json()["rows"]
    for row in rows:
        r = c.patch(
            f"/csf/dimension-scores/{row['id']}",
            headers=h,
            json={
                "governance": 2,
                "policy": 2,
                "implementation": 2,
                "has_evidence": True,
                "target_level": 4,
            },
        )
        assert r.status_code == 200, r.text
    return len(rows)


@pytest.mark.unit
def test_playbook_export_blocked_until_scored_and_approved(app_client) -> None:
    """FIX B-3: the export is a HARD gate. Seed -> export must 409 and name the
    unscored row count (the pre-fix bug shipped a five-artifact deliverable
    asserting Level 1 for every subcategory straight after seeding). Only after
    EVERY in-scope row is scored AND the assessment is approved does it 200, and
    no cell may read "Unscored" or a default "L1"."""
    c, cid = app_client
    r = register_admin_resp(c, "admin@example.com")
    h = {"Authorization": f"Bearer {r.json()['tokens']['access_token']}"}
    svc_id = c.post("/csf/services", headers=h, json={"kind": "nist_csf", "title": "CSF"}).json()[
        "id"
    ]
    assessment = c.post(f"/csf/services/{svc_id}/assessments", headers=h).json()

    # Locked before seeding.
    assert c.post(f"/csf/services/{svc_id}/playbook/export", headers=h).status_code == 409

    c.post(f"/csf/services/{svc_id}/profiles/seed", headers=h, json={"tiers": ["high"]})

    # Seeded but unscored -> 409 whose message names the unscored count.
    blocked = c.post(f"/csf/services/{svc_id}/playbook/export", headers=h)
    assert blocked.status_code == 409, blocked.text
    n = len(SUBCATEGORIES)
    assert str(n) in blocked.json()["error"]["message"], blocked.json()

    # Score every row, but do NOT approve yet -> still 409 (on approval).
    assert _score_all(c, h, svc_id, "high") == n
    pre_approve = c.post(f"/csf/services/{svc_id}/playbook/export", headers=h)
    assert pre_approve.status_code == 409, pre_approve.text
    assert "approved" in pre_approve.json()["error"]["message"].lower()

    # Approve -> export now succeeds.
    c.post(f"/csf/assessments/{assessment['id']}/approve", headers=h)
    ex = c.post(f"/csf/services/{svc_id}/playbook/export", headers=h)
    assert ex.status_code == 200, ex.text
    arts = {a["kind"]: a for a in ex.json()["artifacts"]}
    assert set(arts) == {"xlsx", "exec_pdf", "exec_docx", "full_pdf", "full_docx"}

    dh = {**h, "X-Client-Id": cid}
    magic = {
        "xlsx": b"PK",
        "exec_pdf": b"%PDF-",
        "exec_docx": b"PK",
        "full_pdf": b"%PDF-",
        "full_docx": b"PK",
    }
    for kind, art in arts.items():
        dl = c.get(f"/artifacts/{art['artifact_id']}/download", headers=dh)
        assert dl.status_code == 200, f"{kind}: {dl.status_code}"
        assert dl.content.startswith(magic[kind]), f"{kind} wrong magic bytes"

    # Parse the real XLSX bytes: no data cell may read "Unscored" or a bogus "L1"
    # (every row was scored to L3). The methodology prose mentions "L1 0-2" but
    # never as a bare cell value, so an exact match is safe.
    import io

    from openpyxl import load_workbook

    dl = c.get(f"/artifacts/{arts['xlsx']['artifact_id']}/download", headers=dh)
    wb = load_workbook(io.BytesIO(dl.content))
    seen_level_cell = False
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for val in row:
                assert val != "Unscored", f"{ws.title}: an unscored cell leaked"
                assert val != "L1", f"{ws.title}: a default L1 cell leaked"
                if val == "L3":
                    seen_level_cell = True
    assert seen_level_cell, "expected scored L3 level cells in the workbook"


@pytest.mark.unit
def test_playbook_export_filenames_follow_deliverable_convention(app_client) -> None:
    """FIX B-7: the Playbook export filenames route through the §15.5
    deliverable_filename convention (Company_Service{MMDDYY}[_vN].ext) like every
    other finalize flow, not raw f-strings ("CSF_Playbook_v1.xlsx") with no
    company and no date."""
    from datetime import datetime

    from app.tech_debt.filename import deliverable_filename

    c, cid = app_client
    r = register_admin_resp(c, "admin@example.com")
    h = {"Authorization": f"Bearer {r.json()['tokens']['access_token']}"}
    svc_id = c.post("/csf/services", headers=h, json={"kind": "nist_csf", "title": "CSF"}).json()[
        "id"
    ]
    assessment = c.post(f"/csf/services/{svc_id}/assessments", headers=h).json()
    c.post(f"/csf/services/{svc_id}/profiles/seed", headers=h, json={"tiers": ["high"]})
    _score_all(c, h, svc_id, "high")
    c.post(f"/csf/assessments/{assessment['id']}/approve", headers=h)
    ex = c.post(f"/csf/services/{svc_id}/playbook/export", headers=h)
    assert ex.status_code == 200, ex.text

    filenames = {a["kind"]: a["filename"] for a in ex.json()["artifacts"]}
    today = datetime.now(UTC).date()
    company = "Test Tenant"  # the fixture's tenant legal_name
    expected = {
        "xlsx": deliverable_filename(
            company=company, service_slug="CSF_Playbook", extension="xlsx", day=today, version=1
        ),
        "exec_pdf": deliverable_filename(
            company=company,
            service_slug="CSF_Playbook_Executive",
            extension="pdf",
            day=today,
            version=1,
        ),
        "exec_docx": deliverable_filename(
            company=company,
            service_slug="CSF_Playbook_Executive",
            extension="docx",
            day=today,
            version=1,
        ),
        "full_pdf": deliverable_filename(
            company=company,
            service_slug="CSF_Playbook_Full",
            extension="pdf",
            day=today,
            version=1,
        ),
        "full_docx": deliverable_filename(
            company=company,
            service_slug="CSF_Playbook_Full",
            extension="docx",
            day=today,
            version=1,
        ),
    }
    assert filenames == expected
