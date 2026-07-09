"""Risk Register exporters (Work Order E): XLSX + PDF + Word.

Pure renderers over the register + entries. The XLSX carries every field plus
the blank client-governance columns; the PDF/Word are an executive snapshot
(KPI cards, axis counts, the 5x5 matrix, the tier/cadence legend) plus the full
table. Tool bytes are written by the route layer.
"""

from __future__ import annotations

import io
from collections.abc import Sequence
from dataclasses import dataclass
from typing import Any

from app.risk.engine import (
    IMPACT_ORDER,
    LIKELIHOOD_ORDER,
    Impact,
    Likelihood,
    RecommendedAction,
    RiskAxis,
    RiskTier,
    action_counts,
    axis_counts,
    cadence_for,
    matrix_counts,
    tier_counts,
)

# Blank columns the client uses for governance — SHIELD does not populate these.
_GOVERNANCE_COLUMNS = [
    "Owner",
    "Decision-maker",
    "Approval Date",
    "Acceptance Expiry",
    "Next Review",
    "Status",
]


@dataclass(frozen=True)
class RiskExportContext:
    client_legal_name: str
    version: int
    entries: list[Any]  # RiskEntry rows


def _enum_list(values, enum_cls):
    out = []
    for v in values:
        if not v:
            continue
        try:
            out.append(enum_cls(v))
        except (ValueError, KeyError):
            continue
    return out


def build_context(
    *, client_legal_name: str | None, version: int, entries: Sequence[Any]
) -> RiskExportContext:
    return RiskExportContext(
        client_legal_name=client_legal_name or "Client",
        version=version,
        entries=list(entries),
    )


def _li(e: Any) -> str:
    lk = (e.likelihood or "").replace("_", " ").title()
    im = (e.impact or "").replace("_", " ").title()
    return f"{lk} x {im}".strip(" x")


def _joined(v) -> str:
    return ", ".join(v) if isinstance(v, list) else ""


def _source(e: Any) -> str:
    if e.source and e.source_id:
        return f"{e.source}:{e.source_id}"
    return e.source_id or e.source or ""


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def render_xlsx(ctx: RiskExportContext) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("openpyxl returned no active worksheet")
    ws.title = "Risk Register"

    header = [
        "ID",
        "Weakness",
        "Description",
        "Axis",
        "Source",
        "Linked Techniques",
        "Linked Controls",
        "Likelihood",
        "Impact",
        "Tier",
        "Compensating Controls",
        "Residual Risk",
        "Recommended Action",
        "Rationale",
        "Origin",
        "Trust",
        *_GOVERNANCE_COLUMNS,
    ]
    ws.append(header)
    fill = PatternFill(start_color="FFEEF2F7", end_color="FFEEF2F7", fill_type="solid")
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = fill

    for i, e in enumerate(ctx.entries, start=1):
        ws.append(
            [
                i,
                e.title,
                e.description or "",
                (e.axis or "").title(),
                _source(e),
                _joined(e.linked_techniques),
                _joined(e.linked_controls),
                (e.likelihood or "").replace("_", " ").title(),
                (e.impact or "").replace("_", " ").title(),
                (e.tier or "").title(),
                e.compensating_controls or "",
                e.residual_risk or "",
                (e.recommended_action or "").title(),
                e.rationale or "",
                e.origin,
                e.trust or "",
                # Blank governance columns for the client.
                *["" for _ in _GOVERNANCE_COLUMNS],
            ]
        )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Shared summary used by PDF + DOCX
# ---------------------------------------------------------------------------


def _summary_lines(ctx: RiskExportContext) -> list[str]:
    tiers = _enum_list((e.tier for e in ctx.entries), RiskTier)
    axes = _enum_list((e.axis for e in ctx.entries), RiskAxis)
    actions = _enum_list((e.recommended_action for e in ctx.entries), RecommendedAction)
    tc = tier_counts(tiers)
    ac = axis_counts(axes)
    acts = action_counts(actions)
    crit_high = tc["critical"] + tc["high"]
    return [
        f"Total entries: {len(ctx.entries)}",
        f"Critical + High: {crit_high}",
        f"By axis — detection {ac['detection']}, prevention "
        f"{ac['prevention']}, response {ac['response']}",
        "By recommended action — " + ", ".join(f"{k} {v}" for k, v in acts.items() if v),
    ]


def _legend_rows() -> list[list[str]]:
    return [[t.value.title(), cadence_for(t)] for t in RiskTier]


def _matrix_grid(ctx: RiskExportContext) -> list[list[str]]:
    """The 5x5 Likelihood x Impact count grid (rows = likelihood high->low, cols
    = impact). Shared by the PDF and the DOCX so the two never disagree."""
    matrix = matrix_counts(
        [
            (Likelihood(e.likelihood), Impact(e.impact))
            for e in ctx.entries
            if e.likelihood in Likelihood._value2member_map_
            and e.impact in Impact._value2member_map_
        ]
    )
    grid: list[list[str]] = [[""] + [im.value.title() for im in IMPACT_ORDER]]
    for lk in reversed(LIKELIHOOD_ORDER):
        row = [lk.value.replace("_", " ").title()]
        for im in IMPACT_ORDER:
            cell = next(c for c in matrix if c.likelihood == lk.value and c.impact == im.value)
            row.append(str(cell.count))
        grid.append(row)
    return grid


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def render_pdf(ctx: RiskExportContext) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.7 * inch,
        bottomMargin=0.7 * inch,
        title=f"Risk Register — {ctx.client_legal_name}",
        author="SHIELD by Kentro",
    )
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)
    body = styles["BodyText"]

    def _grid(data, widths):
        t = Table(data, colWidths=widths, repeatRows=1)
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f7")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d6dae3")),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        return t

    story: list = [
        Paragraph(f"Risk Register (v{ctx.version})", h1),
        Paragraph(ctx.client_legal_name, body),
        Spacer(1, 0.2 * inch),
        Paragraph("Summary", h2),
    ]
    for line in _summary_lines(ctx):
        story.append(Paragraph(line, body))

    story.append(Paragraph("Likelihood x Impact matrix", h2))
    story.append(_grid(_matrix_grid(ctx), [1.1 * inch] + [0.9 * inch] * 5))

    story.append(Paragraph("Tier legend (review cadence)", h2))
    story.append(_grid([["Tier", "Suggested cadence"], *_legend_rows()], [1.2 * inch, 5.0 * inch]))

    story.append(PageBreak())
    story.append(Paragraph("Register", h2))
    table = [["ID", "Weakness", "Axis", "L x I", "Tier", "Recommended", "Linked Source"]]
    for i, e in enumerate(ctx.entries, start=1):
        table.append(
            [
                str(i),
                e.title,
                (e.axis or "").title(),
                _li(e),
                (e.tier or "").title(),
                (e.recommended_action or "").title(),
                _source(e),
            ]
        )
    story.append(
        _grid(
            table,
            [0.4 * inch, 2.0 * inch, 0.8 * inch, 1.1 * inch, 0.8 * inch, 1.0 * inch, 1.2 * inch],
        )
    )
    doc.build(story)
    return out.getvalue()


# ---------------------------------------------------------------------------
# DOCX
# ---------------------------------------------------------------------------


def render_docx(ctx: RiskExportContext) -> bytes:
    from app.docx_export import (
        add_heading,
        add_paragraphs,
        add_table,
        add_title,
        new_document,
        to_bytes,
    )

    doc = new_document(f"Risk Register — {ctx.client_legal_name}")
    add_title(doc, f"Risk Register (v{ctx.version})", ctx.client_legal_name)

    add_heading(doc, "Summary")
    add_paragraphs(doc, _summary_lines(ctx))

    # B-7: the module docstring promises "the 5x5 matrix" for "PDF/Word" and the
    # PDF renders it, but the DOCX did not. Port the same grid into Word so the
    # two deliverables agree.
    add_heading(doc, "Likelihood x Impact matrix")
    matrix_grid = _matrix_grid(ctx)
    add_table(doc, matrix_grid[0], matrix_grid[1:])

    add_heading(doc, "Tier legend (review cadence)")
    add_table(doc, ["Tier", "Suggested cadence"], _legend_rows())

    add_heading(doc, "Register")
    rows = [
        [
            str(i),
            e.title,
            (e.axis or "").title(),
            _li(e),
            (e.tier or "").title(),
            (e.recommended_action or "").title(),
            _source(e),
        ]
        for i, e in enumerate(ctx.entries, start=1)
    ]
    add_table(
        doc, ["ID", "Weakness", "Axis", "L x I", "Tier", "Recommended", "Linked Source"], rows
    )

    return to_bytes(doc)
