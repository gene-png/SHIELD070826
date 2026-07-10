"""CSF full-Playbook exports (Work Order D4).

Three deliverables over the data the engine computes (Enterprise weighted-floor
roll-up + per-tier dimension scores + gaps/priorities):

  - render_xlsx           : the data workbook (Enterprise + per-tier sheets).
  - render_exec_*         : an executive briefing (~8-12pp).
  - render_full_*         : the comprehensive playbook (~20-30pp).

The reports use tables with colour-coded maturity-level cells (L1 red -> L5
green) and computed summary sentences (no free-text narrative). All maturity
math is code-computed; the report only presents it.
"""

from __future__ import annotations

import io
from collections.abc import Mapping, Sequence
from html import escape
from typing import Any

# ---------------------------------------------------------------------------
# XLSX workbook
# ---------------------------------------------------------------------------


def _autofit(ws: Any) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(60, max(10, width + 2))


def render_xlsx(
    *,
    client_name: str,
    version: int,
    enterprise_rows: Sequence[Any],
    tier_profiles: Mapping[str, Sequence[Any]],
    action_items: Sequence[Any] = (),
) -> bytes:
    """`enterprise_rows` are EnterpriseSubcategory-like; `tier_profiles` maps a
    tier name to its CsfDimensionScoreResponse-like rows (total/level computed)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    head_fill = PatternFill(start_color="FFEEF2F7", end_color="FFEEF2F7", fill_type="solid")

    def _header(ws: Any, cols: list[str]) -> None:
        ws.append(cols)
        for i in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=i)
            cell.font = Font(bold=True)
            cell.fill = head_fill

    ws = wb.active
    ws.title = "Enterprise Profile"
    _header(
        ws,
        [
            "Subcategory",
            "Function",
            "Outcome",
            "High",
            "Mod",
            "Low",
            "Enterprise",
            "Rule",
            "Target",
            "Gap",
            "Priority",
        ],
    )
    for r in enterprise_rows:
        levels = r.tier_levels
        ws.append(
            [
                r.subcategory_code,
                r.function,
                r.name,
                levels.get("high", ""),
                levels.get("moderate", ""),
                levels.get("low", ""),
                _ent_label(r),
                f"#{r.rollup_rule}",
                f"L{r.target_level}" if r.target_level else "",
                "Yes" if r.gap else "",
                r.priority or "",
            ]
        )
    _autofit(ws)

    for tier in ("high", "moderate", "low"):
        rows = tier_profiles.get(tier)
        if not rows:
            continue
        ts = wb.create_sheet(tier.title())
        _header(
            ts,
            [
                "Subcategory",
                "Governance",
                "Policy & Process",
                "Implementation",
                "Monitoring & Measurement",
                "Continuous Improvement",
                "Total",
                "Level",
                "Evidence capped",
                "In scope",
                "Target",
            ],
        )
        for row in rows:
            ts.append(
                [
                    row.subcategory_code,
                    row.governance,
                    row.policy,
                    row.implementation,
                    row.monitoring,
                    row.improvement,
                    row.total,
                    _tier_label(row),
                    "Yes" if row.evidence_capped else "",
                    "Yes" if row.in_scope else "No",
                    f"L{row.target_level}" if row.target_level else "",
                ]
            )
        _autofit(ts)

    # Action Plan (POA&M) — Playbook Step 10. Structured owner + due-date rows so
    # the action plan ships inside the deliverable instead of a side spreadsheet.
    aps = wb.create_sheet("Action Plan")
    _header(aps, ACTION_PLAN_HEADER)
    action_rows = _action_plan_rows(action_items)
    for row in action_rows:
        aps.append(row)
    if not action_rows:
        aps.append(["—", "No action items recorded", "", "", ""])
    _autofit(aps)

    cover = wb.create_sheet("About", 0)
    cover.append(["SHIELD by Kentro — CSF 2.0 Full Playbook"])
    cover.append([f"Client: {client_name}"])
    cover.append([f"Working profile version: {version}"])
    cover.append([])
    cover.append(
        [
            "Levels are code-computed: total = sum of the five dimensions (0-10); "
            "L1 0-2, L2 3-5, L3 6-7, L4 8-9, L5 10. Enterprise = weighted-floor "
            "roll-up across the tiers in use."
        ]
    )
    cover["A1"].font = Font(bold=True, size=14)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Shared reporting data
# ---------------------------------------------------------------------------

# Maturity level colours: L1 (weakest) red -> L5 (strongest) green.
LEVEL_HEX: dict[int, str] = {
    1: "#fca5a5",
    2: "#fdba74",
    3: "#fde047",
    4: "#bef264",
    5: "#86efac",
}
FUNCTION_NAMES = {
    "GV": "Govern",
    "ID": "Identify",
    "PR": "Protect",
    "DE": "Detect",
    "RS": "Respond",
    "RC": "Recover",
}
FUNCTION_ORDER = ["GV", "ID", "PR", "DE", "RS", "RC"]
_PRIORITY_ORDER = {"P1": 0, "P2": 1, "P3": 2}

METHODOLOGY = [
    "Each subcategory is scored across five dimensions — Governance, Policy & "
    "Process, Implementation, Monitoring & Measurement, and Continuous "
    "Improvement — on a 0-2 scale.",
    "The five dimension scores sum to a 0-10 total, which maps to a maturity "
    "level: Level 1 (0-2), Level 2 (3-5), Level 3 (6-7), Level 4 (8-9), Level "
    "5 (10).",
    "An evidence cap applies: where producible evidence is absent, "
    "Implementation is held to at most 1 and the maturity level to at most 2.",
    "Scoring is performed per FIPS-199 impact tier (HIGH / MODERATE / LOW) as "
    "separate working profiles, then rolled up to a single Enterprise level per "
    "subcategory using a weighted-floor rule set so that weak high-impact "
    "systems are not masked by stronger low-impact ones.",
    "AI assists by drafting the dimension scores; all totals, levels, the "
    "evidence cap, the roll-up, gaps, and priorities are computed "
    "deterministically in code.",
]


# FIX B-3 layer 3 (belt and braces). Rows carry a `scored` flag (default True for
# any caller that doesn't set one). A row that was never scored renders
# "Unscored" for its maturity level instead of an all-zero "L1", and is left
# unshaded. The export gate already blocks unscored rows, so these paths are a
# defensive backstop, never the normal case.
_UNSCORED = "Unscored"


def _is_scored(r: Any) -> bool:
    return getattr(r, "scored", True) is not False


def _ent_label(r: Any) -> str:
    return f"L{r.enterprise_level}" if _is_scored(r) else _UNSCORED


def _tier_label(r: Any) -> str:
    return f"L{r.level}" if _is_scored(r) else _UNSCORED


def _ent_shades(rows: Sequence[Any]) -> list[int | None]:
    return [(r.enterprise_level if _is_scored(r) else None) for r in rows]


def _fn_code(raw: Any) -> str:
    return str(raw or "").split(".")[-1].upper()


def _overall_level(rows: Sequence[Any]) -> int:
    levels = [r.enterprise_level for r in rows]
    return round(sum(levels) / len(levels)) if levels else 0


def _priority_counts(rows: Sequence[Any]) -> dict[str, int]:
    return {p: sum(1 for r in rows if r.priority == p) for p in ("P1", "P2", "P3")}


def _function_summary(rows: Sequence[Any]) -> list[dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = {}
    for r in rows:
        code = _fn_code(r.function)
        g = groups.setdefault(code, {"levels": [], "gaps": 0, "targets": []})
        g["levels"].append(r.enterprise_level)
        if r.gap:
            g["gaps"] += 1
        if r.target_level:
            g["targets"].append(r.target_level)
    ordered = [c for c in FUNCTION_ORDER if c in groups] + [
        c for c in groups if c not in FUNCTION_ORDER
    ]
    out: list[dict[str, Any]] = []
    for code in ordered:
        g = groups[code]
        avg = round(sum(g["levels"]) / len(g["levels"])) if g["levels"] else 0
        out.append(
            {
                "code": code,
                "name": FUNCTION_NAMES.get(code, code or "Other"),
                "count": len(g["levels"]),
                "avg": avg,
                "gaps": g["gaps"],
                "target": max(g["targets"]) if g["targets"] else None,
            }
        )
    return out


def _overview_sentences(rows: Sequence[Any]) -> list[str]:
    total = len(rows)
    overall = _overall_level(rows)
    gaps = sum(1 for r in rows if r.gap)
    pc = _priority_counts(rows)
    fns = _function_summary(rows)
    lines = [
        f"This assessment covers {total} in-scope NIST CSF 2.0 subcategories. "
        f"Enterprise maturity, rolled up across the impact tiers in use, "
        f"averages Level {overall} of 5.",
        f"{gaps} subcategories fall short of their target maturity — "
        f"{pc['P1']} Priority 1 (critical), {pc['P2']} Priority 2, and "
        f"{pc['P3']} Priority 3.",
    ]
    if len(fns) >= 2:
        strongest = max(fns, key=lambda f: f["avg"])
        weakest = min(fns, key=lambda f: f["avg"])
        if strongest["code"] != weakest["code"]:
            lines.append(
                f"The strongest function is {strongest['name']} "
                f"(avg Level {strongest['avg']}); the area needing the most "
                f"attention is {weakest['name']} (avg Level {weakest['avg']})."
            )
    return lines


def _next_steps(rows: Sequence[Any]) -> list[str]:
    pc = _priority_counts(rows)
    steps: list[str] = []
    if pc["P1"]:
        steps.append(
            f"Remediate the {pc['P1']} Priority 1 gap(s) first — these are "
            f"Core-metric, high-impact, multi-system weaknesses."
        )
    if pc["P2"]:
        steps.append(f"Schedule the {pc['P2']} Priority 2 gap(s) into the next planning cycle.")
    if pc["P3"]:
        steps.append(f"Track the {pc['P3']} Priority 3 gap(s) for continuous improvement.")
    if not steps:
        steps.append(
            "No gaps were identified — maintain current controls and re-assess on the next cycle."
        )
    return steps


def _gap_rows(rows: Sequence[Any], *, limit: int | None = None) -> list[Any]:
    gaps = [r for r in rows if r.gap]
    gaps.sort(key=lambda r: (_PRIORITY_ORDER.get(r.priority or "P3", 3), r.subcategory_code))
    return gaps[:limit] if limit else gaps


# ---------------------------------------------------------------------------
# Action plan / POA&M (Playbook Step 10, FIX H-8)
# ---------------------------------------------------------------------------

ACTION_PLAN_HEADER = ["Subcategory", "Owner", "Due date", "Milestone", "Status"]
_ACTION_STATUS_LABELS = {"open": "Open", "in_progress": "In progress", "done": "Done"}


def _action_status_label(item: Any) -> str:
    raw = str(getattr(item, "status", "") or "")
    return _ACTION_STATUS_LABELS.get(raw, raw)


def _action_due(item: Any) -> str:
    d = getattr(item, "due_date", None)
    return d.isoformat() if d else ""


def _action_plan_rows(action_items: Sequence[Any]) -> list[list[str]]:
    """One [subcategory, owner, due date, milestone, status] row per action item.

    Owner defaults to "Unassigned" so a committed but not-yet-owned action still
    reads as a plan, not a blank. Ordering is the caller's (query orders by
    subcategory then created_at)."""
    return [
        [
            it.subcategory_code,
            it.owner or "Unassigned",
            _action_due(it),
            (getattr(it, "milestone", None) or ""),
            _action_status_label(it),
        ]
        for it in action_items
    ]


# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------


def _pdf_styles() -> dict[str, Any]:
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

    base = getSampleStyleSheet()
    return {
        "title": base["Title"],
        "body": base["BodyText"],
        "h2": ParagraphStyle("h2", parent=base["Heading2"], spaceBefore=14, spaceAfter=6),
        "cell": ParagraphStyle("cell", fontSize=8, leading=10),
        "small": ParagraphStyle("small", parent=base["BodyText"], fontSize=8, textColor="#64748b"),
    }


def _pdf_table(
    header: list[str],
    body: list[list[Any]],
    widths: list[float],
    *,
    color_col: int | None = None,
    color_levels: Sequence[int | None] | None = None,
) -> Any:
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    data = [header, *body]
    style: list[Any] = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d6dae3")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]
    if color_col is not None and color_levels is not None:
        for i, lvl in enumerate(color_levels, start=1):
            if lvl:
                style.append(
                    ("BACKGROUND", (color_col, i), (color_col, i), colors.HexColor(LEVEL_HEX[lvl]))
                )
    t = Table(data, colWidths=widths, repeatRows=1)
    t.setStyle(TableStyle(style))
    return t


def _cover(
    story: list[Any],
    styles: dict[str, Any],
    *,
    subtitle: str,
    client_name: str,
    version: int,
    generated_on: str | None,
) -> None:
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, Spacer

    story.append(Spacer(1, 1.5 * inch))
    story.append(Paragraph("NIST CSF 2.0", styles["title"]))
    story.append(Paragraph(subtitle, styles["h2"]))
    story.append(Spacer(1, 0.4 * inch))
    story.append(Paragraph(f"Prepared for: <b>{escape(client_name)}</b>", styles["body"]))
    story.append(Paragraph(f"Working profile version: {version}", styles["body"]))
    if generated_on:
        story.append(Paragraph(f"Generated: {escape(generated_on)}", styles["body"]))
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph("Prepared by SHIELD by Kentro.", styles["body"]))
    story.append(
        Paragraph(
            "Confidential — contains a security assessment of the named "
            "organization. Distribute on a need-to-know basis.",
            styles["small"],
        )
    )


def _scorecard(story: list[Any], styles: dict[str, Any], rows: Sequence[Any]) -> None:
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph

    story.append(Paragraph("Maturity scorecard", styles["h2"]))
    fns = _function_summary(rows)
    body = [
        [
            f["name"],
            str(f["count"]),
            f"L{f['avg']}",
            f"L{f['target']}" if f["target"] else "—",
            str(f["gaps"]),
        ]
        for f in fns
    ]
    levels = [f["avg"] for f in fns]
    story.append(
        _pdf_table(
            ["Function", "Subcategories", "Maturity", "Target", "Gaps"],
            body,
            [2.2 * inch, 1.3 * inch, 1.1 * inch, 1.0 * inch, 0.9 * inch],
            color_col=2,
            color_levels=levels,
        )
    )


def _gap_table(story: list[Any], styles: dict[str, Any], gaps: Sequence[Any]) -> None:
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph

    if not gaps:
        story.append(
            Paragraph("No gaps — every in-scope subcategory meets its target.", styles["body"])
        )
        return
    body = [
        [
            g.subcategory_code,
            Paragraph(escape(g.name), styles["cell"]),
            _ent_label(g),
            f"L{g.target_level}" if g.target_level else "—",
            g.priority or "",
        ]
        for g in gaps
    ]
    levels = _ent_shades(gaps)
    story.append(
        _pdf_table(
            ["Subcategory", "Outcome", "Current", "Target", "Priority"],
            body,
            [0.9 * inch, 3.0 * inch, 0.8 * inch, 0.8 * inch, 0.8 * inch],
            color_col=2,
            color_levels=levels,
        )
    )


def _action_plan_table_pdf(
    story: list[Any], styles: dict[str, Any], action_items: Sequence[Any]
) -> None:
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph

    story.append(
        Paragraph(
            "Owners and target dates for closing each gap. Maintained in-platform "
            "so the plan travels with the deliverable.",
            styles["body"],
        )
    )
    rows = _action_plan_rows(action_items)
    if not rows:
        story.append(Paragraph("No action items recorded for this assessment yet.", styles["body"]))
        return
    body = [
        [
            r[0],
            r[1],
            r[2] or "—",
            Paragraph(escape(r[3]), styles["cell"]),
            r[4],
        ]
        for r in rows
    ]
    story.append(
        _pdf_table(
            ACTION_PLAN_HEADER,
            body,
            [0.9 * inch, 1.4 * inch, 1.0 * inch, 2.6 * inch, 1.0 * inch],
        )
    )


def _new_pdf(out: io.BytesIO, title: str, client_name: str) -> Any:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate

    return SimpleDocTemplate(
        out,
        pagesize=letter,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=0.7 * inch,
        bottomMargin=0.7 * inch,
        title=f"{title} — {client_name}",
        author="SHIELD by Kentro",
    )


def render_exec_pdf(
    *,
    client_name: str,
    version: int,
    enterprise_rows: Sequence[Any],
    generated_on: str | None = None,
) -> bytes:
    from reportlab.platypus import PageBreak, Paragraph

    out = io.BytesIO()
    doc = _new_pdf(out, "CSF 2.0 Executive Briefing", client_name)
    styles = _pdf_styles()
    story: list[Any] = []
    _cover(
        story,
        styles,
        subtitle="Executive Briefing",
        client_name=client_name,
        version=version,
        generated_on=generated_on,
    )
    story.append(PageBreak())

    story.append(Paragraph("Executive summary", styles["h2"]))
    for line in _overview_sentences(enterprise_rows):
        story.append(Paragraph(line, styles["body"]))
    _scorecard(story, styles, enterprise_rows)
    story.append(Paragraph("Top priority gaps", styles["h2"]))
    _gap_table(story, styles, _gap_rows(enterprise_rows, limit=12))
    story.append(Paragraph("Recommended next steps", styles["h2"]))
    for step in _next_steps(enterprise_rows):
        story.append(Paragraph(f"• {step}", styles["body"]))
    doc.build(story)
    return out.getvalue()


def render_full_pdf(
    *,
    client_name: str,
    version: int,
    enterprise_rows: Sequence[Any],
    generated_on: str | None = None,
    action_items: Sequence[Any] = (),
) -> bytes:
    from reportlab.lib.units import inch
    from reportlab.platypus import PageBreak, Paragraph

    out = io.BytesIO()
    doc = _new_pdf(out, "CSF 2.0 Full Playbook", client_name)
    styles = _pdf_styles()
    story: list[Any] = []
    _cover(
        story,
        styles,
        subtitle="Full Playbook",
        client_name=client_name,
        version=version,
        generated_on=generated_on,
    )
    story.append(PageBreak())

    story.append(Paragraph("Contents", styles["h2"]))
    for n, sec in enumerate(
        [
            "Executive summary",
            "Methodology",
            "Maturity scorecard",
            "Function detail",
            "Prioritized roadmap",
            "Action plan (POA&M)",
            "Appendix — all subcategories",
        ],
        start=1,
    ):
        story.append(Paragraph(f"{n}. {sec}", styles["body"]))
    story.append(PageBreak())

    story.append(Paragraph("1. Executive summary", styles["h2"]))
    for line in _overview_sentences(enterprise_rows):
        story.append(Paragraph(line, styles["body"]))

    story.append(Paragraph("2. Methodology", styles["h2"]))
    for line in METHODOLOGY:
        story.append(Paragraph(line, styles["body"]))

    story.append(Paragraph("3. Maturity scorecard", styles["h2"]))
    _scorecard(story, styles, enterprise_rows)

    story.append(PageBreak())
    story.append(Paragraph("4. Function detail", styles["h2"]))
    by_fn: dict[str, list[Any]] = {}
    for r in enterprise_rows:
        by_fn.setdefault(_fn_code(r.function), []).append(r)
    ordered = [c for c in FUNCTION_ORDER if c in by_fn] + [
        c for c in by_fn if c not in FUNCTION_ORDER
    ]
    for code in ordered:
        frows = sorted(by_fn[code], key=lambda r: r.subcategory_code)
        name = FUNCTION_NAMES.get(code, code)
        avg = round(sum(r.enterprise_level for r in frows) / len(frows))
        fgaps = sum(1 for r in frows if r.gap)
        story.append(Paragraph(f"{name} ({code})", styles["h2"]))
        story.append(
            Paragraph(
                f"{len(frows)} subcategories · average Level {avg} · {fgaps} gap(s).",
                styles["body"],
            )
        )
        body = [
            [
                r.subcategory_code,
                Paragraph(escape(r.name), styles["cell"]),
                _ent_label(r),
                f"L{r.target_level}" if r.target_level else "—",
                ("Yes" if r.gap else ""),
                r.priority or "",
            ]
            for r in frows
        ]
        story.append(
            _pdf_table(
                ["Subcategory", "Outcome", "Maturity", "Target", "Gap", "Priority"],
                body,
                [0.9 * inch, 2.6 * inch, 0.85 * inch, 0.8 * inch, 0.6 * inch, 0.75 * inch],
                color_col=2,
                color_levels=_ent_shades(frows),
            )
        )

    story.append(PageBreak())
    story.append(Paragraph("5. Prioritized roadmap", styles["h2"]))
    _gap_table(story, styles, _gap_rows(enterprise_rows))

    story.append(PageBreak())
    story.append(Paragraph("6. Action plan (POA&M)", styles["h2"]))
    _action_plan_table_pdf(story, styles, action_items)

    story.append(PageBreak())
    story.append(Paragraph("7. Appendix — all subcategories", styles["h2"]))
    appx = [
        [
            r.subcategory_code,
            FUNCTION_NAMES.get(_fn_code(r.function), _fn_code(r.function)),
            r.tier_levels.get("high", "—"),
            r.tier_levels.get("moderate", "—"),
            r.tier_levels.get("low", "—"),
            _ent_label(r),
            f"#{r.rollup_rule}",
            ("Yes" if r.gap else ""),
        ]
        for r in sorted(enterprise_rows, key=lambda r: r.subcategory_code)
    ]
    story.append(
        _pdf_table(
            ["Subcat", "Function", "H", "M", "L", "Ent.", "Rule", "Gap"],
            appx,
            [
                0.9 * inch,
                1.2 * inch,
                0.5 * inch,
                0.5 * inch,
                0.5 * inch,
                0.6 * inch,
                0.5 * inch,
                0.5 * inch,
            ],
            color_col=5,
            color_levels=_ent_shades(sorted(enterprise_rows, key=lambda r: r.subcategory_code)),
        )
    )
    doc.build(story)
    return out.getvalue()


# ---------------------------------------------------------------------------
# Word (.docx) reports
# ---------------------------------------------------------------------------


def _shade_col(table: Any, col: int, levels: Sequence[int | None]) -> None:
    from app.docx_export import shade_cell

    for i, lvl in enumerate(levels):
        if lvl:
            shade_cell(table.rows[i + 1].cells[col], LEVEL_HEX[lvl])


def _docx_cover(
    doc: Any, *, subtitle: str, client_name: str, version: int, generated_on: str | None
) -> None:
    from app.docx_export import add_paragraphs, add_title

    add_title(doc, "NIST CSF 2.0", subtitle)
    meta = [f"Prepared for: {client_name}", f"Working profile version: {version}"]
    if generated_on:
        meta.append(f"Generated: {generated_on}")
    meta.append("Prepared by SHIELD by Kentro.")
    meta.append(
        "Confidential — contains a security assessment of the named "
        "organization. Distribute on a need-to-know basis."
    )
    add_paragraphs(doc, meta)


def _docx_scorecard(doc: Any, rows: Sequence[Any]) -> None:
    from app.docx_export import add_heading, add_table

    add_heading(doc, "Maturity scorecard")
    fns = _function_summary(rows)
    table = add_table(
        doc,
        ["Function", "Subcategories", "Maturity", "Target", "Gaps"],
        [
            [
                f["name"],
                f["count"],
                f"L{f['avg']}",
                f"L{f['target']}" if f["target"] else "—",
                f["gaps"],
            ]
            for f in fns
        ],
    )
    _shade_col(table, 2, [f["avg"] for f in fns])


def render_exec_docx(
    *,
    client_name: str,
    version: int,
    enterprise_rows: Sequence[Any],
    generated_on: str | None = None,
) -> bytes:
    from app.docx_export import add_heading, add_paragraphs, add_table, new_document, to_bytes

    doc = new_document(f"CSF 2.0 Executive Briefing — {client_name}")
    _docx_cover(
        doc,
        subtitle="Executive Briefing",
        client_name=client_name,
        version=version,
        generated_on=generated_on,
    )

    add_heading(doc, "Executive summary")
    add_paragraphs(doc, _overview_sentences(enterprise_rows))
    _docx_scorecard(doc, enterprise_rows)

    add_heading(doc, "Top priority gaps")
    gaps = _gap_rows(enterprise_rows, limit=12)
    if gaps:
        table = add_table(
            doc,
            ["Subcategory", "Outcome", "Current", "Target", "Priority"],
            [
                [
                    g.subcategory_code,
                    g.name,
                    _ent_label(g),
                    f"L{g.target_level}" if g.target_level else "—",
                    g.priority or "",
                ]
                for g in gaps
            ],
        )
        _shade_col(table, 2, _ent_shades(gaps))
    else:
        add_paragraphs(doc, ["No gaps — every in-scope subcategory meets its target."])

    add_heading(doc, "Recommended next steps")
    add_paragraphs(doc, [f"• {s}" for s in _next_steps(enterprise_rows)])
    return to_bytes(doc)


def render_full_docx(
    *,
    client_name: str,
    version: int,
    enterprise_rows: Sequence[Any],
    generated_on: str | None = None,
    action_items: Sequence[Any] = (),
) -> bytes:
    from app.docx_export import (
        add_heading,
        add_page_break,
        add_paragraphs,
        add_table,
        new_document,
        to_bytes,
    )

    doc = new_document(f"CSF 2.0 Full Playbook — {client_name}")
    _docx_cover(
        doc,
        subtitle="Full Playbook",
        client_name=client_name,
        version=version,
        generated_on=generated_on,
    )
    add_page_break(doc)

    add_heading(doc, "1. Executive summary")
    add_paragraphs(doc, _overview_sentences(enterprise_rows))

    add_heading(doc, "2. Methodology")
    add_paragraphs(doc, METHODOLOGY)

    add_heading(doc, "3. Maturity scorecard")
    _docx_scorecard(doc, enterprise_rows)

    add_page_break(doc)
    add_heading(doc, "4. Function detail")
    by_fn: dict[str, list[Any]] = {}
    for r in enterprise_rows:
        by_fn.setdefault(_fn_code(r.function), []).append(r)
    ordered = [c for c in FUNCTION_ORDER if c in by_fn] + [
        c for c in by_fn if c not in FUNCTION_ORDER
    ]
    for code in ordered:
        frows = sorted(by_fn[code], key=lambda r: r.subcategory_code)
        name = FUNCTION_NAMES.get(code, code)
        avg = round(sum(r.enterprise_level for r in frows) / len(frows))
        fgaps = sum(1 for r in frows if r.gap)
        add_heading(doc, f"{name} ({code})", level=2)
        add_paragraphs(doc, [f"{len(frows)} subcategories · average Level {avg} · {fgaps} gap(s)."])
        table = add_table(
            doc,
            ["Subcategory", "Outcome", "Maturity", "Target", "Gap", "Priority"],
            [
                [
                    r.subcategory_code,
                    r.name,
                    _ent_label(r),
                    f"L{r.target_level}" if r.target_level else "—",
                    ("Yes" if r.gap else ""),
                    r.priority or "",
                ]
                for r in frows
            ],
        )
        _shade_col(table, 2, _ent_shades(frows))

    add_page_break(doc)
    add_heading(doc, "5. Prioritized roadmap")
    gaps = _gap_rows(enterprise_rows)
    if gaps:
        table = add_table(
            doc,
            ["Subcategory", "Outcome", "Current", "Target", "Priority"],
            [
                [
                    g.subcategory_code,
                    g.name,
                    _ent_label(g),
                    f"L{g.target_level}" if g.target_level else "—",
                    g.priority or "",
                ]
                for g in gaps
            ],
        )
        _shade_col(table, 2, _ent_shades(gaps))
    else:
        add_paragraphs(doc, ["No gaps — every in-scope subcategory meets its target."])

    add_page_break(doc)
    add_heading(doc, "6. Action plan (POA&M)")
    add_paragraphs(
        doc,
        [
            "Owners and target dates for closing each gap. Maintained in-platform "
            "so the plan travels with the deliverable."
        ],
    )
    action_rows = _action_plan_rows(action_items)
    if action_rows:
        add_table(doc, ACTION_PLAN_HEADER, action_rows)
    else:
        add_paragraphs(doc, ["No action items recorded for this assessment yet."])

    add_page_break(doc)
    add_heading(doc, "7. Appendix — all subcategories")
    ordered_rows = sorted(enterprise_rows, key=lambda r: r.subcategory_code)
    table = add_table(
        doc,
        ["Subcat", "Function", "High", "Mod", "Low", "Enterprise", "Rule", "Gap"],
        [
            [
                r.subcategory_code,
                FUNCTION_NAMES.get(_fn_code(r.function), _fn_code(r.function)),
                r.tier_levels.get("high", "—"),
                r.tier_levels.get("moderate", "—"),
                r.tier_levels.get("low", "—"),
                _ent_label(r),
                f"#{r.rollup_rule}",
                ("Yes" if r.gap else ""),
            ]
            for r in ordered_rows
        ],
    )
    _shade_col(table, 5, _ent_shades(ordered_rows))
    return to_bytes(doc)
