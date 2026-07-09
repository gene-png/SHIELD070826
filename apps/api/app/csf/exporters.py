"""CSF 2.0 deliverable renderers - turn an assessment into XLSX + PDF bytes.

Master Spec §15 Phase 4: each released service produces a PDF + XLSX
deliverable. Reuses the structural pattern from
`app.tech_debt.exporters` (reportlab + openpyxl, pure functions, no
DB/IO). The route layer writes the bytes via the existing
StorageBackend abstraction.

XLSX sheets:
  - "Score Summary": overall + per-function maturity rollup
  - "Answers":       per-subcategory tier + notes
  - "Gap Plan":      prioritized remediation gaps

PDF:
  Executive summary page with overall maturity + per-function bars,
  followed by the top-N gap table.
"""

from __future__ import annotations

import io
from collections.abc import Iterable
from dataclasses import dataclass
from typing import TYPE_CHECKING

from app.csf.catalog import FUNCTIONS, SUBCATEGORIES, FunctionCode, Subcategory
from app.csf.gap import GapAnalysis
from app.csf.maturity import tier_label
from app.csf.scoring import ScoreResult

if TYPE_CHECKING:
    from reportlab.platypus import TableStyle
from app.models.csf_assessment import CsfAnswer, CsfAssessment

# The PDF/DOCX narratives print only the top slice; the XLSX Gap Plan sheet
# carries the full list. The heading states "Top N of <total_gap_count>" so the
# narrative can never silently contradict the true gap count (B-4).
NARRATIVE_GAP_LIMIT = 20


@dataclass(frozen=True)
class CsfDeliverableContext:
    """Inputs the renderers share. Built once by the route layer."""

    client_legal_name: str
    service_title: str
    assessment: CsfAssessment
    answers: list[CsfAnswer]
    score: ScoreResult
    gap: GapAnalysis


def build_context(
    *,
    client_legal_name: str | None,
    service_title: str,
    assessment: CsfAssessment,
    answers: Iterable[CsfAnswer],
    score: ScoreResult,
    gap: GapAnalysis,
) -> CsfDeliverableContext:
    return CsfDeliverableContext(
        client_legal_name=client_legal_name or "Client",
        service_title=service_title,
        assessment=assessment,
        answers=list(answers),
        score=score,
        gap=gap,
    )


def _function_name(code: FunctionCode) -> str:
    for f in FUNCTIONS:
        if f.code == code:
            return f.name
    return code.value


def _subcategory_meta(code: str) -> Subcategory | None:
    for s in SUBCATEGORIES:
        if s.code == code:
            return s
    return None


def _fmt_tier(value: float | None) -> str:
    if value is None:
        return "—"
    return f"{value:.2f}"


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def render_xlsx(ctx: CsfDeliverableContext) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Remove the default blank sheet; we add three named ones below.
    default = wb.active
    if default is not None:
        wb.remove(default)

    header_fill = PatternFill(start_color="FFEEF2F7", end_color="FFEEF2F7", fill_type="solid")
    bold = Font(bold=True)
    italic = Font(italic=True)

    # --- Sheet 1: Score Summary ---
    ws = wb.create_sheet("Score Summary")
    ws.append(["Engagement", ctx.client_legal_name])
    ws.append(["Service", ctx.service_title])
    ws.append(["Assessment version", ctx.assessment.version])
    ws.append(["Overall maturity", ctx.score.overall_maturity_label])
    ws.append(["Average tier", _fmt_tier(ctx.score.average_tier)])
    ws.append(["Coverage", f"{ctx.score.answered_subcategories}/{ctx.score.total_subcategories}"])
    for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold
    ws.append([])
    ws.append(["Function", "Name", "Answered", "Total", "Coverage %", "Average tier"])
    for col_idx in range(1, 7):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = bold
        cell.fill = header_fill
    for fs in ctx.score.by_function:
        ws.append(
            [
                fs.function.value,
                fs.function_name,
                fs.answered_count,
                fs.subcategory_count,
                fs.coverage_pct,
                _fmt_tier(fs.average_tier),
            ]
        )
    for w, col in zip([10, 28, 12, 10, 14, 16], range(1, 7), strict=True):
        ws.column_dimensions[get_column_letter(col)].width = w

    # --- Sheet 2: Answers ---
    ws2 = wb.create_sheet("Answers")
    headers = [
        "Subcategory",
        "Function",
        "Category",
        "Name",
        "Outcome",
        "Tier",
        "Tier label",
        "Notes",
    ]
    ws2.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    answers_by_code = {a.subcategory_code: a for a in ctx.answers}
    # Iterate the canonical catalog order so missing answers still
    # render as blank rows (the assessor sees what wasn't scored).
    for sc in SUBCATEGORIES:
        ans = answers_by_code.get(sc.code)
        tier = ans.maturity_tier if ans else None
        notes = ans.notes if ans else None
        ws2.append(
            [
                sc.code,
                sc.function.value,
                sc.category,
                sc.name,
                sc.outcome,
                tier if tier is not None else "",
                tier_label(tier) if tier is not None else "Unscored",
                notes or "",
            ]
        )
    for w, col in zip([14, 10, 10, 32, 60, 8, 16, 60], range(1, 9), strict=True):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # --- Sheet 3: Gap Plan ---
    ws3 = wb.create_sheet("Gap Plan")
    headers3 = [
        "Subcategory",
        "Function",
        "Category",
        "Name",
        "Current tier",
        "Target tier",
        "Gap size",
        "Priority",
        "Notes",
    ]
    ws3.append(headers3)
    for col in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=col)
        cell.font = bold
        cell.fill = header_fill
    for g in ctx.gap.gaps:
        ws3.append(
            [
                g.code,
                g.function.value,
                g.category,
                g.name,
                g.current_tier,
                g.target_tier,
                g.gap_size,
                g.priority_score,
                g.notes or "",
            ]
        )
    if not ctx.gap.gaps:
        ws3.append(
            [
                "—",
                "",
                "",
                "No gaps at target tier",
                "",
                ctx.gap.target_tier,
                0,
                0,
                "",
            ]
        )
        ws3.cell(row=2, column=4).font = italic
    for w, col in zip([14, 10, 10, 32, 14, 14, 12, 12, 50], range(1, 10), strict=True):
        ws3.column_dimensions[get_column_letter(col)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def render_docx(ctx: CsfDeliverableContext) -> bytes:
    """Word deliverable mirroring the PDF (Work Order C4)."""
    from app.docx_export import (
        add_heading,
        add_paragraphs,
        add_table,
        add_title,
        new_document,
        to_bytes,
    )

    doc = new_document(f"{ctx.service_title} — {ctx.client_legal_name}")
    add_title(doc, ctx.service_title, ctx.client_legal_name)

    add_heading(doc, "Maturity summary")
    add_paragraphs(
        doc,
        [
            f"Overall maturity: {ctx.score.overall_maturity_label}",
            f"Average tier: {_fmt_tier(ctx.score.average_tier)}",
            f"Coverage: {ctx.score.answered_subcategories}/"
            f"{ctx.score.total_subcategories} ({ctx.score.coverage_pct}%)",
        ],
    )

    add_heading(doc, "Per-function rollup")
    add_table(
        doc,
        ["Function", "Name", "Average tier", "Coverage"],
        [
            [
                fs.function.value,
                fs.function_name,
                _fmt_tier(fs.average_tier),
                f"{fs.answered_count}/{fs.subcategory_count} ({fs.coverage_pct}%)",
            ]
            for fs in ctx.score.by_function
        ],
    )

    shown = ctx.gap.gaps[:NARRATIVE_GAP_LIMIT]
    total = ctx.gap.total_gap_count
    add_heading(
        doc,
        f"Top {len(shown)} of {total} remediation gaps (target T{ctx.gap.target_tier})",
    )
    if not shown:
        add_paragraphs(
            doc,
            [f"No gaps at target tier {ctx.gap.target_tier} " f"({ctx.gap.target_label})."],
        )
    else:
        if total > len(shown):
            add_paragraphs(
                doc,
                [
                    "See the Gap Plan sheet of the accompanying XLSX workbook for the "
                    f"full list of all {total} remediation gaps.",
                ],
            )
        add_table(
            doc,
            ["Code", "Function", "Subcategory", "Current → Target", "Priority"],
            [
                [
                    g.code,
                    g.function.value,
                    g.name,
                    f"T{g.current_tier} → T{g.target_tier}",
                    f"{g.priority_score:.2f}",
                ]
                for g in shown
            ],
        )

    return to_bytes(doc)


def render_pdf(ctx: CsfDeliverableContext) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
    )

    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.7 * inch,
        bottomMargin=0.7 * inch,
        title=f"{ctx.service_title} — {ctx.client_legal_name}",
        author="SHIELD by Kentro",
    )
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)
    body = styles["BodyText"]

    story: list = []
    story.append(Paragraph(ctx.service_title, h1))
    story.append(Paragraph(ctx.client_legal_name, body))
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("Maturity summary", h2))
    story.append(
        Paragraph(
            f"Overall maturity: <b>{ctx.score.overall_maturity_label}</b> · "
            f"Average tier: <b>{_fmt_tier(ctx.score.average_tier)}</b> · "
            f"Coverage: <b>{ctx.score.answered_subcategories}/"
            f"{ctx.score.total_subcategories}</b> "
            f"({ctx.score.coverage_pct}%)",
            body,
        )
    )

    story.append(Paragraph("Per-function rollup", h2))
    fn_table_data: list[list] = [["Function", "Name", "Average tier", "Coverage"]]
    for fs in ctx.score.by_function:
        fn_table_data.append(
            [
                fs.function.value,
                fs.function_name,
                _fmt_tier(fs.average_tier),
                f"{fs.answered_count}/{fs.subcategory_count} ({fs.coverage_pct}%)",
            ]
        )
    fn_table = Table(
        fn_table_data,
        colWidths=[0.8 * inch, 2.2 * inch, 1.4 * inch, 2.0 * inch],
        repeatRows=1,
    )
    fn_table.setStyle(_table_style())
    story.append(fn_table)

    story.append(PageBreak())

    shown = ctx.gap.gaps[:NARRATIVE_GAP_LIMIT]
    total = ctx.gap.total_gap_count
    story.append(
        Paragraph(
            f"Top {len(shown)} of {total} remediation gaps (target T{ctx.gap.target_tier})",
            h2,
        )
    )
    if not shown:
        story.append(
            Paragraph(
                f"No gaps at target tier {ctx.gap.target_tier} " f"({ctx.gap.target_label}).",
                body,
            )
        )
    else:
        if total > len(shown):
            story.append(
                Paragraph(
                    "See the Gap Plan sheet of the accompanying XLSX workbook for the "
                    f"full list of all {total} remediation gaps.",
                    body,
                )
            )
        gap_table_data: list[list] = [
            ["Code", "Function", "Subcategory", "Current → Target", "Priority"]
        ]
        for g in shown:
            gap_table_data.append(
                [
                    g.code,
                    g.function.value,
                    g.name,
                    f"T{g.current_tier} → T{g.target_tier}",
                    f"{g.priority_score:.2f}",
                ]
            )
        gap_table = Table(
            gap_table_data,
            colWidths=[0.9 * inch, 0.8 * inch, 3.0 * inch, 1.4 * inch, 0.9 * inch],
            repeatRows=1,
        )
        gap_table.setStyle(_table_style())
        story.append(gap_table)

    doc.build(story)
    return out.getvalue()


def _table_style() -> TableStyle:
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle

    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f7")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0e1220")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d6dae3")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ]
    )


__all__ = [
    "CsfDeliverableContext",
    "build_context",
    "render_pdf",
    "render_xlsx",
]
