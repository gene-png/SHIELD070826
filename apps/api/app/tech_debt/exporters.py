"""Deliverable renderers - turn a capability list into XLSX + PDF bytes.

Master Spec §15 Phase 3: "PDF + XLSX exporters for the deliverable."

XLSX: openpyxl. Header row + one row per capability + a summary row at
the bottom (Total Cost, Estimated Savings).

PDF: ReportLab. Pure Python; no native deps required (unlike WeasyPrint).
Phase 6 polish can revisit visual fidelity, but for v1 the deliverable is
a real, legitimate PDF with a title, summary, table, and savings figure.

Both renderers are pure functions over the data; no DB, no I/O. The
route layer writes the bytes via the existing StorageBackend.
"""

from __future__ import annotations

import io
from collections.abc import Iterable
from dataclasses import dataclass

from app.models.capability import CapabilityDisposition, CapabilityItem, CapabilityList


@dataclass(frozen=True)
class DeliverableContext:
    """Inputs the renderers share. Built once by the route layer."""

    client_legal_name: str
    service_title: str
    cap_list: CapabilityList
    items: list[CapabilityItem]
    total_cost: float
    estimated_savings: float
    savings_cost_known: bool


def _disposition_label(d: CapabilityDisposition | None) -> str:
    if d is None:
        return "Undecided"
    return {
        CapabilityDisposition.KEEP: "Keep",
        CapabilityDisposition.CONSOLIDATE: "Consolidate",
        CapabilityDisposition.CUT: "Cut",
    }[d]


def build_context(
    *,
    client_legal_name: str | None,
    service_title: str,
    cap_list: CapabilityList,
    items: Iterable[CapabilityItem],
) -> DeliverableContext:
    items_list = list(items)
    total_cost = 0.0
    estimated_savings = 0.0
    savings_known = True
    for it in items_list:
        if it.annual_cost_usd is not None:
            total_cost += float(it.annual_cost_usd)
        if it.disposition == CapabilityDisposition.CUT:
            if it.annual_cost_usd is None:
                savings_known = False
            else:
                estimated_savings += float(it.annual_cost_usd)
    return DeliverableContext(
        client_legal_name=client_legal_name or "Client",
        service_title=service_title,
        cap_list=cap_list,
        items=items_list,
        total_cost=total_cost,
        estimated_savings=estimated_savings,
        savings_cost_known=savings_known,
    )


# ---------------------------------------------------------------------------
# Shared analysis - the numbers behind every deliverable format.
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class DeliverableAnalysis:
    """Aggregates derived once, shared by the HTML / DOCX / XLSX renderers."""

    disposition_counts: dict[str, int]  # keep / consolidate / cut / undecided
    category_count: int
    spend_by_category: list[tuple[str, float]]  # descending, all categories
    overlap: object  # app.tech_debt.overlap.OverlapAnalysis
    savings_str: str
    savings_pct: str


def _build_analysis(ctx: DeliverableContext) -> DeliverableAnalysis:
    from collections import defaultdict

    from .overlap import analyze_overlap

    counts = {"keep": 0, "consolidate": 0, "cut": 0, "undecided": 0}
    for it in ctx.items:
        counts[_disposition_slug(it.disposition)] += 1

    cat_spend: dict[str, float] = defaultdict(float)
    for it in ctx.items:
        if it.annual_cost_usd is not None:
            cat = (it.category or "Uncategorized").strip() or "Uncategorized"
            cat_spend[cat] += float(it.annual_cost_usd)
    spend_by_category = sorted(cat_spend.items(), key=lambda kv: kv[1], reverse=True)

    savings_str = (
        _fmt_usd(ctx.estimated_savings)
        if ctx.savings_cost_known
        else f"≥ {_fmt_usd(ctx.estimated_savings)}"
    )
    savings_pct = (
        f"{(ctx.estimated_savings / ctx.total_cost * 100):.1f}% annual reduction"
        if ctx.total_cost > 0
        else "via consolidation"
    )
    category_count = len(
        {(i.category or "").strip() for i in ctx.items if (i.category or "").strip()}
    )

    return DeliverableAnalysis(
        disposition_counts=counts,
        category_count=category_count,
        spend_by_category=spend_by_category,
        overlap=analyze_overlap(ctx.items),
        savings_str=savings_str,
        savings_pct=savings_pct,
    )


# ---------------------------------------------------------------------------
# XLSX - inventory plus the full analysis, one sheet per view.
# ---------------------------------------------------------------------------


def render_xlsx(ctx: DeliverableContext) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    an = _build_analysis(ctx)
    overlap = an.overlap
    header_fill = PatternFill(start_color="FFEEF2F7", end_color="FFEEF2F7", fill_type="solid")

    def _style_header(ws, ncols: int) -> None:
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def _widths(ws, widths: list[int]) -> None:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()

    # --- Sheet 1: Inventory ---
    ws = wb.active
    if ws is None:
        raise RuntimeError("openpyxl returned no active worksheet")
    ws.title = "Inventory"
    header = [
        "Name",
        "Vendor",
        "Category",
        "Function",
        "Annual Cost (USD)",
        "Licenses",
        "Disposition",
        "Rationale",
        "Notes",
        "AI Confidence %",
    ]
    ws.append(header)
    _style_header(ws, len(header))
    for item in ctx.items:
        ws.append(
            [
                item.name,
                item.vendor or "",
                item.category or "",
                item.function or "",
                float(item.annual_cost_usd) if item.annual_cost_usd is not None else "",
                item.license_count if item.license_count is not None else "",
                _disposition_label(item.disposition),
                item.disposition_rationale or "",
                item.notes or "",
                item.confidence_pct if item.confidence_pct is not None else "",
            ]
        )
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=5).number_format = "$#,##0"
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="Total annual cost").font = Font(bold=True)
    ws.cell(row=summary_row, column=5, value=ctx.total_cost).number_format = "$#,##0"
    ws.cell(row=summary_row + 1, column=1, value="Estimated annual savings").font = Font(bold=True)
    sc = ws.cell(row=summary_row + 1, column=5, value=ctx.estimated_savings)
    sc.number_format = "$#,##0"
    if not ctx.savings_cost_known:
        ws.cell(
            row=summary_row + 1, column=6, value="≥ (one or more cut rows missing a cost)"
        ).font = Font(italic=True)
    _widths(ws, [28, 22, 16, 28, 18, 10, 14, 38, 38, 16])

    # --- Sheet 2: Spend by Category ---
    ws2 = wb.create_sheet("Spend by Category")
    ws2.append(["Category", "Annual Spend (USD)", "Tool Count", "% of Total"])
    _style_header(ws2, 4)
    tool_counts: dict[str, int] = {}
    for it in ctx.items:
        c = (it.category or "Uncategorized").strip() or "Uncategorized"
        tool_counts[c] = tool_counts.get(c, 0) + 1
    for cat, spend in an.spend_by_category:
        ws2.append(
            [
                cat,
                spend,
                tool_counts.get(cat, 0),
                (spend / ctx.total_cost) if ctx.total_cost else 0,
            ]
        )
    for r in range(2, ws2.max_row + 1):
        ws2.cell(row=r, column=2).number_format = "$#,##0"
        ws2.cell(row=r, column=4).number_format = "0.0%"
    _widths(ws2, [28, 20, 12, 12])

    # --- Sheet 3: Overlaps ---
    ws3 = wb.create_sheet("Overlaps")
    ws3.append(["Type", "Group", "Tool Count", "Combined Cost (USD)", "Cost Known", "Tools"])
    _style_header(ws3, 6)
    for b in overlap.by_category:
        ws3.append(
            [
                "Category",
                b.key,
                b.item_count,
                b.total_cost,
                "Yes" if b.cost_known else "No",
                ", ".join(b.item_names),
            ]
        )
    for b in overlap.by_vendor:
        ws3.append(
            [
                "Vendor",
                b.key,
                b.item_count,
                b.total_cost,
                "Yes" if b.cost_known else "No",
                ", ".join(b.item_names),
            ]
        )
    for r in range(2, ws3.max_row + 1):
        ws3.cell(row=r, column=4).number_format = "$#,##0"
    _widths(ws3, [12, 24, 12, 20, 12, 60])

    # --- Sheet 4: Consolidation Plan ---
    ws4 = wb.create_sheet("Consolidation Plan")
    ws4.append(["Metric", "Value"])
    _style_header(ws4, 2)
    c = an.disposition_counts
    plan_rows = [
        ("Capabilities reviewed", len(ctx.items)),
        ("Functional categories", an.category_count),
        ("Overlap clusters (category)", len(overlap.by_category)),
        ("Keep", c["keep"]),
        ("Consolidate", c["consolidate"]),
        ("Cut", c["cut"]),
        ("Undecided", c["undecided"]),
        ("Total annual cost", ctx.total_cost),
        ("Estimated annual savings", ctx.estimated_savings),
    ]
    for label, val in plan_rows:
        ws4.append([label, val])
    ws4.cell(row=ws4.max_row - 1, column=2).number_format = "$#,##0"
    ws4.cell(row=ws4.max_row, column=2).number_format = "$#,##0"
    if not ctx.savings_cost_known:
        ws4.append(["Note", "Savings is a lower bound - one or more Cut rows lack a cost."])
    _widths(ws4, [32, 24])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def render_pdf(ctx: DeliverableContext) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.7 * inch,
        bottomMargin=0.7 * inch,
        title=f"{ctx.service_title} — {ctx.client_legal_name}",
        author="SHIELD by Kentro",
    )
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)
    body = styles["BodyText"]

    story: list = []
    story.append(Paragraph(ctx.service_title, h1))
    story.append(Paragraph(ctx.client_legal_name, body))
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("Summary", h2))
    savings = (
        f"${ctx.estimated_savings:,.0f}"
        if ctx.savings_cost_known
        else f"≥ ${ctx.estimated_savings:,.0f}"
    )
    story.append(
        Paragraph(
            f"Capabilities reviewed: <b>{len(ctx.items)}</b> · "
            f"Total annual cost: <b>${ctx.total_cost:,.0f}</b> · "
            f"Estimated annual savings: <b>{savings}</b>",
            body,
        )
    )
    if not ctx.savings_cost_known:
        story.append(
            Paragraph(
                "Note: at least one row marked <i>Cut</i> is missing an annual cost. "
                "The savings figure is a lower bound.",
                body,
            )
        )

    story.append(Paragraph("Capability list", h2))

    table_data: list[list] = [["Name", "Vendor", "Category", "Annual cost", "Disposition"]]
    for item in ctx.items:
        cost = f"${float(item.annual_cost_usd):,.0f}" if item.annual_cost_usd is not None else "—"
        table_data.append(
            [
                item.name,
                item.vendor or "",
                item.category or "",
                cost,
                _disposition_label(item.disposition),
            ]
        )

    table = Table(
        table_data,
        colWidths=[2.2 * inch, 1.4 * inch, 1.2 * inch, 1.0 * inch, 1.2 * inch],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f7")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0e1220")),
                ("ALIGN", (3, 1), (3, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d6dae3")),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ]
        )
    )
    story.append(table)

    doc.build(story)
    return out.getvalue()


# ---------------------------------------------------------------------------
# DOCX (Work Order C4) - mirrors the PDF content.
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# HTML dashboard (primary, auto-built, in-browser view)
# ---------------------------------------------------------------------------


def _fmt_usd(value: float | None, *, dash: str = "—") -> str:
    if value is None:
        return dash
    return f"${value:,.0f}"


def _disposition_slug(d: CapabilityDisposition | None) -> str:
    if d is None:
        return "undecided"
    return {
        CapabilityDisposition.KEEP: "keep",
        CapabilityDisposition.CONSOLIDATE: "consolidate",
        CapabilityDisposition.CUT: "cut",
    }[d]


def render_html_dashboard(ctx: DeliverableContext) -> bytes:
    """Self-contained executive dashboard (dark theme, no external assets).

    Charts are pure CSS bars rather than a JS charting CDN, so the file
    renders identically offline and inside the FedRAMP boundary with no
    egress. Every dynamic string is HTML-escaped (values originate from AI
    extraction + client uploads).
    """
    from collections import defaultdict
    from html import escape

    from .overlap import analyze_overlap

    items = ctx.items
    overlap = analyze_overlap(items)

    # Disposition tallies for the consolidation summary.
    counts = {"keep": 0, "consolidate": 0, "cut": 0, "undecided": 0}
    for it in items:
        counts[_disposition_slug(it.disposition)] += 1

    # Spend by category (top 10), for the horizontal bar chart.
    cat_spend: dict[str, float] = defaultdict(float)
    for it in items:
        cat = (it.category or "Uncategorized").strip() or "Uncategorized"
        if it.annual_cost_usd is not None:
            cat_spend[cat] += float(it.annual_cost_usd)
    top_cats = sorted(cat_spend.items(), key=lambda kv: kv[1], reverse=True)[:10]
    max_cat = max((v for _, v in top_cats), default=0.0)

    max_vendor = max((b.item_count for b in overlap.by_vendor), default=0)

    savings_str = (
        _fmt_usd(ctx.estimated_savings)
        if ctx.savings_cost_known
        else f"≥ {_fmt_usd(ctx.estimated_savings)}"
    )
    savings_pct = (
        f"{(ctx.estimated_savings / ctx.total_cost * 100):.1f}% annual reduction"
        if ctx.total_cost > 0
        else "via consolidation"
    )
    initials = "".join(w[0] for w in ctx.client_legal_name.split()[:2]).upper() or "SH"

    # ---- KPI cards ----
    kpis = f"""
    <div class="kpi">
      <div class="label">Applications</div>
      <div class="value">{len(items)}</div>
      <div class="sub">Across {len({(i.category or '').strip() for i in items if (i.category or '').strip()})} functional categories</div>
    </div>
    <div class="kpi amber">
      <div class="label">Annual License Spend</div>
      <div class="value">{_fmt_usd(ctx.total_cost, dash='$0')}</div>
      <div class="sub">{overlap.no_cost_count} item(s) missing a cost</div>
    </div>
    <div class="kpi red">
      <div class="label">Overlap Clusters</div>
      <div class="value">{len(overlap.by_category)}</div>
      <div class="sub">Categories with more than one tool</div>
    </div>
    <div class="kpi green">
      <div class="label">Estimated Savings</div>
      <div class="value">{escape(savings_str)}</div>
      <div class="sub">{escape(savings_pct)}</div>
    </div>
    """

    # ---- Spend-by-category bars ----
    if top_cats:
        cat_bars = "".join(f"""
        <div class="bar-row">
          <div class="bar-label" title="{escape(cat)}">{escape(cat)}</div>
          <div class="bar-track"><div class="bar-fill" style="width:{(spend / max_cat * 100) if max_cat else 0:.1f}%"></div></div>
          <div class="bar-value">{_fmt_usd(spend)}</div>
        </div>""" for cat, spend in top_cats)
    else:
        cat_bars = '<p class="desc">No costed items to chart.</p>'

    # ---- Vendor sprawl bars ----
    if overlap.by_vendor:
        vendor_bars = "".join(f"""
        <div class="bar-row">
          <div class="bar-label" title="{escape(b.key)}">{escape(b.key)}</div>
          <div class="bar-track"><div class="bar-fill cyan" style="width:{(b.item_count / max_vendor * 100) if max_vendor else 0:.1f}%"></div></div>
          <div class="bar-value">{b.item_count} tools</div>
        </div>""" for b in overlap.by_vendor[:10])
    else:
        vendor_bars = '<p class="desc">No vendor appears more than once.</p>'

    # ---- Overlap clusters (category buckets) ----
    if overlap.by_category:
        clusters = "".join(f"""
        <div class="rcard {'high' if b.item_count >= 3 else 'medium'}">
          <div class="rhead">
            <span class="cat">{escape(b.key)}</span>
            <span class="sev">{b.item_count} tools</span>
          </div>
          <div class="products">
            {''.join(f'<span class="product-tag">{escape(n)}</span>' for n in b.item_names)}
          </div>
          <div class="reco">Combined spend: <b>{_fmt_usd(b.total_cost) if b.cost_known else '≥ ' + _fmt_usd(b.total_cost)}</b> — consolidation candidate.</div>
        </div>""" for b in overlap.by_category)
    else:
        clusters = '<p class="desc">No functional overlaps detected — every category has a single tool.</p>'

    # ---- Consolidation disposition chips ----
    disp_summary = f"""
      <div class="debt-list">
        <div class="debt"><div class="num">K</div><h3>Keep — {counts['keep']}</h3><p>Tools retained as the primary capability in their category.</p></div>
        <div class="debt"><div class="num">C</div><h3>Consolidate — {counts['consolidate']}</h3><p>Overlapping tools to fold into a primary or renegotiate.</p></div>
        <div class="debt"><div class="num">X</div><h3>Cut — {counts['cut']}</h3><p>Redundant or unused tools flagged for retirement.</p><span class="tag">{savings_str} projected</span></div>
      </div>
    """

    # ---- Inventory table ----
    rows = "".join(f"""
        <tr class="{'flag' if (it.confidence_pct is not None and it.confidence_pct < 70) else ''}">
          <td>{escape(it.name)}</td>
          <td>{escape(it.vendor or '—')}</td>
          <td><span class="cat-chip">{escape((it.category or 'Uncategorized').strip() or 'Uncategorized')}</span></td>
          <td>{escape(it.function or '—')}</td>
          <td class="cost {'free' if it.annual_cost_usd is None else ''}">{_fmt_usd(float(it.annual_cost_usd) if it.annual_cost_usd is not None else None)}</td>
          <td>{it.license_count if it.license_count is not None else '—'}</td>
          <td><span class="disp {_disposition_slug(it.disposition)}">{_disposition_label(it.disposition)}</span></td>
          <td>{f'{it.confidence_pct}%' if it.confidence_pct is not None else '—'}</td>
        </tr>""" for it in items)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1.0" />
<title>{escape(ctx.client_legal_name)} — {escape(ctx.service_title)}</title>
<style>
  :root {{
    --bg:#0b1020; --bg-2:#11172d; --panel:#151b35; --panel-2:#1a2143; --border:#232a4d;
    --text:#e6e9f5; --muted:#98a2c4; --accent:#6366f1; --accent-2:#22d3ee;
    --green:#10b981; --amber:#f59e0b; --red:#ef4444;
  }}
  * {{ box-sizing:border-box; }}
  html,body {{ margin:0; padding:0; background:
      radial-gradient(1200px 600px at 0% 0%, #1a2046 0%, var(--bg) 50%) fixed,
      radial-gradient(900px 600px at 100% 0%, #1a1d3d 0%, transparent 60%) fixed, var(--bg);
    color:var(--text); font-family:-apple-system,BlinkMacSystemFont,"Inter","Segoe UI",Roboto,sans-serif;
    min-height:100vh; -webkit-font-smoothing:antialiased; }}
  .wrap {{ max-width:1440px; margin:0 auto; padding:32px 28px 64px; }}
  .header {{ display:flex; align-items:center; justify-content:space-between; margin-bottom:28px; flex-wrap:wrap; gap:16px; }}
  .brand {{ display:flex; align-items:center; gap:14px; }}
  .logo {{ width:44px; height:44px; border-radius:12px; background:linear-gradient(135deg,var(--accent),var(--accent-2));
    display:flex; align-items:center; justify-content:center; font-weight:800; color:#fff; font-size:16px; box-shadow:0 8px 24px rgba(99,102,241,.35); }}
  .brand h1 {{ margin:0; font-size:20px; font-weight:700; letter-spacing:-.01em; }}
  .brand p {{ margin:2px 0 0; font-size:13px; color:var(--muted); }}
  .badge {{ display:inline-flex; align-items:center; gap:6px; background:rgba(99,102,241,.12); color:#c7d2fe;
    border:1px solid rgba(99,102,241,.35); padding:6px 12px; border-radius:999px; font-size:12px; font-weight:600; }}
  .badge .dot {{ width:8px; height:8px; border-radius:50%; background:var(--green); box-shadow:0 0 0 4px rgba(16,185,129,.18); }}
  .kpis {{ display:grid; grid-template-columns:repeat(4,1fr); gap:16px; margin-bottom:22px; }}
  .kpi {{ background:linear-gradient(180deg,var(--panel) 0%,var(--panel-2) 100%); border:1px solid var(--border);
    border-radius:16px; padding:20px 22px; position:relative; overflow:hidden; }}
  .kpi::after {{ content:""; position:absolute; right:-30px; top:-30px; width:120px; height:120px; border-radius:50%;
    opacity:.14; background:radial-gradient(circle,var(--accent),transparent 70%); }}
  .kpi.amber::after {{ background:radial-gradient(circle,var(--amber),transparent 70%); }}
  .kpi.red::after {{ background:radial-gradient(circle,var(--red),transparent 70%); }}
  .kpi.green::after {{ background:radial-gradient(circle,var(--green),transparent 70%); }}
  .kpi .label {{ font-size:12px; color:var(--muted); text-transform:uppercase; letter-spacing:.08em; font-weight:600; }}
  .kpi .value {{ font-size:30px; font-weight:700; margin-top:8px; letter-spacing:-.02em; }}
  .kpi .sub {{ font-size:12px; color:var(--muted); margin-top:6px; }}
  .charts-row {{ display:grid; grid-template-columns:1fr 1fr; gap:18px; margin-bottom:18px; }}
  .section {{ background:linear-gradient(180deg,var(--panel) 0%,var(--panel-2) 100%); border:1px solid var(--border);
    border-radius:16px; padding:22px; margin-bottom:18px; }}
  .section h2 {{ margin:0 0 4px; font-size:16px; font-weight:700; display:flex; align-items:center; gap:10px; }}
  .section h2 .pill {{ font-size:11px; padding:3px 8px; border-radius:999px; background:rgba(99,102,241,.18); color:#c7d2fe; font-weight:600; }}
  .section .desc {{ color:var(--muted); font-size:13px; margin:0 0 16px; }}
  .bar-row {{ display:grid; grid-template-columns:150px 1fr 110px; align-items:center; gap:12px; margin-bottom:10px; }}
  .bar-label {{ font-size:12.5px; color:var(--text); white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
  .bar-track {{ height:14px; background:var(--bg-2); border-radius:999px; overflow:hidden; }}
  .bar-fill {{ height:100%; border-radius:999px; background:linear-gradient(90deg,var(--accent),var(--accent-2)); }}
  .bar-fill.cyan {{ background:linear-gradient(90deg,var(--accent-2),var(--green)); }}
  .bar-value {{ font-size:12px; color:var(--muted); text-align:right; font-variant-numeric:tabular-nums; }}
  .redundancy-grid {{ display:grid; grid-template-columns:repeat(2,1fr); gap:14px; }}
  .rcard {{ background:var(--bg-2); border:1px solid var(--border); border-left:3px solid var(--amber); border-radius:12px;
    padding:16px 18px; display:flex; flex-direction:column; gap:8px; }}
  .rcard.high {{ border-left-color:var(--red); }}
  .rcard.medium {{ border-left-color:var(--amber); }}
  .rcard .rhead {{ display:flex; align-items:center; justify-content:space-between; }}
  .rcard .cat {{ font-weight:700; font-size:14px; }}
  .rcard .sev {{ font-size:10px; font-weight:700; text-transform:uppercase; letter-spacing:.08em; padding:3px 8px; border-radius:999px;
    background:rgba(245,158,11,.18); color:#fde68a; }}
  .rcard.high .sev {{ background:rgba(239,68,68,.18); color:#fecaca; }}
  .products {{ display:flex; flex-wrap:wrap; gap:6px; }}
  .product-tag {{ background:rgba(255,255,255,.04); border:1px solid var(--border); color:var(--text); font-size:11.5px;
    padding:4px 9px; border-radius:6px; }}
  .reco {{ font-size:12.5px; color:var(--muted); line-height:1.55; }}
  .reco b {{ color:var(--text); font-weight:600; }}
  .debt-list {{ display:grid; grid-template-columns:repeat(3,1fr); gap:14px; }}
  .debt {{ background:var(--bg-2); border:1px solid var(--border); border-radius:12px; padding:16px 18px; }}
  .debt .num {{ width:28px; height:28px; border-radius:8px; background:linear-gradient(135deg,var(--accent),var(--accent-2));
    display:inline-flex; align-items:center; justify-content:center; font-weight:800; font-size:13px; color:#fff; margin-bottom:10px; }}
  .debt h3 {{ margin:0 0 6px; font-size:14px; font-weight:700; }}
  .debt p {{ margin:0; color:var(--muted); font-size:12.5px; line-height:1.55; }}
  .debt .tag {{ display:inline-block; margin-top:10px; font-size:11px; color:#a7f3d0; background:rgba(16,185,129,.15); padding:3px 8px; border-radius:999px; }}
  .table-scroll {{ max-height:520px; overflow-y:auto; border:1px solid var(--border); border-radius:10px; }}
  table {{ width:100%; border-collapse:collapse; font-size:12.5px; }}
  thead th {{ position:sticky; top:0; background:var(--bg-2); color:var(--muted); text-align:left; padding:10px 12px;
    font-weight:600; text-transform:uppercase; letter-spacing:.06em; font-size:11px; border-bottom:1px solid var(--border); }}
  tbody td {{ padding:10px 12px; border-bottom:1px solid rgba(255,255,255,.04); }}
  tbody tr:hover {{ background:rgba(99,102,241,.04); }}
  tbody tr.flag td:first-child {{ border-left:2px solid var(--amber); }}
  .cat-chip {{ display:inline-block; padding:2px 8px; border-radius:999px; font-size:11px; font-weight:600; background:rgba(99,102,241,.15); color:#c7d2fe; }}
  .cost {{ font-variant-numeric:tabular-nums; font-weight:600; }}
  .cost.free {{ color:var(--muted); font-weight:500; }}
  .disp {{ font-size:11px; font-weight:700; padding:2px 8px; border-radius:999px; }}
  .disp.keep {{ background:rgba(16,185,129,.15); color:#a7f3d0; }}
  .disp.consolidate {{ background:rgba(245,158,11,.15); color:#fde68a; }}
  .disp.cut {{ background:rgba(239,68,68,.15); color:#fecaca; }}
  .disp.undecided {{ background:rgba(152,162,196,.15); color:var(--muted); }}
  footer {{ color:var(--muted); font-size:11.5px; text-align:center; margin-top:24px; }}
  @media (max-width:1080px) {{ .kpis{{grid-template-columns:repeat(2,1fr);}} .charts-row{{grid-template-columns:1fr;}}
    .redundancy-grid{{grid-template-columns:1fr;}} .debt-list{{grid-template-columns:1fr;}} .bar-row{{grid-template-columns:120px 1fr 90px;}} }}
</style>
</head>
<body>
<div class="wrap">
  <div class="header">
    <div class="brand">
      <div class="logo">{escape(initials)}</div>
      <div>
        <h1>{escape(ctx.client_legal_name)} — Software Portfolio</h1>
        <p>Executive Dashboard · {escape(ctx.service_title)}</p>
      </div>
    </div>
    <div class="badge"><span class="dot"></span> SHIELD by Kentro · Capability list v{ctx.cap_list.version}</div>
  </div>

  <div class="kpis">{kpis}</div>

  <div class="charts-row">
    <div class="section">
      <h2>Annual Spend by Category <span class="pill">Top {len(top_cats)}</span></h2>
      <p class="desc">Where the money goes across the reviewed portfolio.</p>
      {cat_bars}
    </div>
    <div class="section">
      <h2>Tool Sprawl by Vendor</h2>
      <p class="desc">Vendors supplying more than one tool — license-negotiation and consolidation candidates.</p>
      {vendor_bars}
    </div>
  </div>

  <div class="section">
    <h2>Functional Overlaps <span class="pill">{len(overlap.by_category)} clusters</span></h2>
    <p class="desc">Categories with more than one tool. Each cluster is a consolidation candidate.</p>
    <div class="redundancy-grid">{clusters}</div>
  </div>

  <div class="section">
    <h2>Consolidation Plan</h2>
    <p class="desc">Disposition of every reviewed capability.</p>
    {disp_summary}
  </div>

  <div class="section">
    <h2>Full Inventory <span class="pill">{len(items)} items</span></h2>
    <p class="desc">Every extracted capability. Amber-flagged rows had AI confidence below 70% and were reviewed.</p>
    <div class="table-scroll">
      <table>
        <thead><tr>
          <th>Name</th><th>Vendor</th><th>Category</th><th>Function</th>
          <th>Annual cost</th><th>Licenses</th><th>Disposition</th><th>AI conf.</th>
        </tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div>
  </div>

  <footer>Generated by SHIELD by Kentro · {escape(ctx.service_title)} · Total annual cost {_fmt_usd(ctx.total_cost, dash='$0')} · Estimated savings {escape(savings_str)}</footer>
</div>
</body>
</html>""".encode()


# ---------------------------------------------------------------------------
# DOCX (Work Order C4) - mirrors the PDF content.
# ---------------------------------------------------------------------------


def render_docx(ctx: DeliverableContext) -> bytes:
    """Executive-view Word report - the narrative counterpart to the HTML
    dashboard: KPI summary, spend by category, functional overlaps,
    consolidation plan, then the full inventory."""
    from app.docx_export import (
        add_heading,
        add_paragraphs,
        add_table,
        add_title,
        new_document,
        shade_cell,
        to_bytes,
    )

    an = _build_analysis(ctx)
    overlap = an.overlap
    c = an.disposition_counts
    ACCENT = "#1a2143"

    def _shade_header(table) -> None:
        for cell in table.rows[0].cells:
            shade_cell(cell, ACCENT)

    doc = new_document(f"{ctx.service_title} — {ctx.client_legal_name}")
    add_title(doc, f"{ctx.client_legal_name} — Software Portfolio", ctx.service_title)

    # --- Executive summary KPIs ---
    add_heading(doc, "Executive Summary")
    add_paragraphs(
        doc,
        [
            f"This review covers {len(ctx.items)} software capabilities across "
            f"{an.category_count} functional categories, representing "
            f"{_fmt_usd(ctx.total_cost, dash='$0')} in annual license spend. "
            f"{len(overlap.by_category)} categories contain more than one tool and are "
            f"consolidation candidates. Estimated annual savings from the current plan: "
            f"{an.savings_str} ({an.savings_pct}).",
        ],
    )
    kpi = add_table(
        doc,
        ["Metric", "Value"],
        [
            ["Applications reviewed", str(len(ctx.items))],
            ["Annual license spend", _fmt_usd(ctx.total_cost, dash="$0")],
            ["Overlap clusters", str(len(overlap.by_category))],
            ["Estimated annual savings", an.savings_str],
        ],
    )
    _shade_header(kpi)

    # --- Spend by category (top 10) ---
    add_heading(doc, "Annual Spend by Category")
    tool_counts: dict[str, int] = {}
    for it in ctx.items:
        cat = (it.category or "Uncategorized").strip() or "Uncategorized"
        tool_counts[cat] = tool_counts.get(cat, 0) + 1
    spend_tbl = add_table(
        doc,
        ["Category", "Annual spend", "Tools"],
        [
            [cat, _fmt_usd(spend), str(tool_counts.get(cat, 0))]
            for cat, spend in an.spend_by_category[:10]
        ],
    )
    _shade_header(spend_tbl)

    # --- Functional overlaps ---
    add_heading(doc, "Functional Overlaps")
    if overlap.by_category:
        ov_tbl = add_table(
            doc,
            ["Category", "Tools", "Combined cost", "Products"],
            [
                [
                    b.key,
                    str(b.item_count),
                    (_fmt_usd(b.total_cost) if b.cost_known else f"≥ {_fmt_usd(b.total_cost)}"),
                    ", ".join(b.item_names),
                ]
                for b in overlap.by_category
            ],
        )
        _shade_header(ov_tbl)
    else:
        add_paragraphs(doc, ["No functional overlaps detected — every category has a single tool."])

    # --- Consolidation plan ---
    add_heading(doc, "Consolidation Plan")
    plan_tbl = add_table(
        doc,
        ["Disposition", "Count", "Meaning"],
        [
            ["Keep", str(c["keep"]), "Primary tool retained in its category"],
            ["Consolidate", str(c["consolidate"]), "Fold into a primary or renegotiate"],
            ["Cut", str(c["cut"]), "Redundant / unused — flagged for retirement"],
            ["Undecided", str(c["undecided"]), "Pending consultant disposition"],
        ],
    )
    _shade_header(plan_tbl)
    if not ctx.savings_cost_known:
        add_paragraphs(
            doc,
            [
                "Note: at least one row marked Cut is missing an annual cost; the savings figure is a lower bound."
            ],
        )

    # --- Full inventory ---
    add_heading(doc, "Full Inventory")
    inv = add_table(
        doc,
        ["Name", "Vendor", "Category", "Annual cost", "Disposition"],
        [
            [
                item.name,
                item.vendor or "",
                item.category or "",
                _fmt_usd(float(item.annual_cost_usd) if item.annual_cost_usd is not None else None),
                _disposition_label(item.disposition),
            ]
            for item in ctx.items
        ],
    )
    _shade_header(inv)

    return to_bytes(doc)
