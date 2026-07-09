"""Inventory file parsers - CSV and XLSX.

Master Spec §15 Phase 3: "Capability list ingest (Excel upload + AI
extraction with redaction)". This module turns the raw artifact bytes
into a row-shaped representation the LLM can reason about. The LLM does
the column-mapping; we just give it well-shaped rows.

Phase 3 only supports CSV + XLSX. PDF ingest is a Phase 6 hardening
target (table extraction is a different problem and the inventory
documents Eugene's customers ship are reliably tabular).

FIX C-3: previously only ``wb.active`` was parsed (garbage from a cover
sheet), duplicate headers silently collapsed (dict-keyed rows, last
column wins), and cells beyond the header width were dropped. We now
scan ALL worksheets and pick the best candidate, uniquify duplicate
headers (``name`` / ``name_2``), keep overflow cells under generated
headers, and return sheet metadata so the caller can tell the admin
which sheet was read, how many rows parsed / were skipped, and whether
truncation occurred.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable
from dataclasses import dataclass
from typing import Any


class UnsupportedInventoryFormat(ValueError):
    """Raised when an artifact's MIME isn't a recognized inventory format."""


class EmptyInventoryError(ValueError):
    """Raised when an inventory file yields zero *data* rows to extract.

    Header-only sheets, blank files, or anything that parses down to no
    usable rows land here. Calling the LLM with nothing to extract is both
    a waste and the road to fabricated output, so the route turns this into
    a typed 422 *before* any provider call.
    """


class CorruptInventoryError(ValueError):
    """Raised when a file advertised as .xlsx can't be opened as one.

    openpyxl raises low-level ``zipfile.BadZipFile`` /
    ``openpyxl ... InvalidFileException`` for OLE2 legacy .xls bytes or a
    truncated/corrupt .xlsx. We convert those at the parse boundary so the
    route can return a typed 422 instead of a raw 500.
    """


# MIME types that the ingest endpoint accepts. Legacy .xls
# (``application/vnd.ms-excel``) is intentionally absent: openpyxl cannot
# read OLE2 .xls, so those are rejected at upload with an actionable message
# (re-save as .xlsx) rather than crashing extraction.
SUPPORTED_MIME = {
    "text/csv": "csv",
    "text/plain": "csv",  # treat .txt as CSV; most inventory exports save this way
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
}

# Max rows we ship to the LLM. Above this and we'd either bust the context
# window or pay a fortune in tokens. v1 inventories are typically 50-300 rows.
MAX_ROWS = 500

# Internal marker key appended by parse_inventory when the input exceeds
# MAX_ROWS. It carries the "truncated" signal but is NOT a data row: it must
# never be shipped to the LLM or persisted as a capability.
SENTINEL_KEY = "__truncated__"


@dataclass(frozen=True)
class ParsedInventory:
    """Parsed rows plus the metadata the caller reports back to the admin.

    ``rows`` still carries the truncation sentinel as its last element when
    the input exceeded MAX_ROWS, so ``data_rows`` / ``was_truncated`` keep
    working on it unchanged (C-3 preserves that contract).
    """

    rows: list[dict[str, Any]]
    sheet_name: str | None
    sheet_count: int
    rows_parsed: int
    rows_skipped: int
    truncated: bool


def is_sentinel_row(row: Any) -> bool:
    """True for the internal truncation marker (see SENTINEL_KEY)."""
    return isinstance(row, dict) and SENTINEL_KEY in row


def data_rows(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    """The real data rows, with any truncation sentinel skipped."""
    return [row for row in rows if not is_sentinel_row(row)]


def was_truncated(rows: Iterable[dict[str, Any]]) -> bool:
    """Whether the parsed rows carry a truncation sentinel."""
    return any(is_sentinel_row(row) for row in rows)


def kind_for_mime(mime_type: str) -> str:
    try:
        return SUPPORTED_MIME[mime_type]
    except KeyError as exc:
        raise UnsupportedInventoryFormat(
            f"Inventory format {mime_type!r} not supported. Accept CSV or XLSX."
        ) from exc


def _uniquify(names: list[str]) -> list[str]:
    """Return `names` with duplicates disambiguated: `name`, `name_2`, ...

    Blank names become a generated `colN` first, so a header row with two
    empty columns doesn't collapse either. Matches the "last column wins"
    bug's inverse: every source column survives as its own key.
    """
    seen: dict[str, int] = {}
    out: list[str] = []
    for i, raw in enumerate(names):
        base = raw.strip() if raw and raw.strip() else f"col{i + 1}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        out.append(base if count == 1 else f"{base}_{count}")
    return out


def _grid_to_rows(grid: list[list[str]]) -> tuple[list[dict[str, Any]], int, list[str]]:
    """Turn a 2-D string grid into row-dicts.

    The first non-empty row is the header. Overflow cells (a data row wider
    than the header) are kept under generated `colN` headers rather than
    dropped. Returns (data_rows, skipped_count, header_names).
    """
    # Find the header: first row with any non-empty cell.
    header_idx = None
    for i, row in enumerate(grid):
        if any(cell.strip() for cell in row):
            header_idx = i
            break
    if header_idx is None:
        return [], 0, []

    header_cells = grid[header_idx]
    body = grid[header_idx + 1 :]

    # Effective width spans the header AND the widest data row so overflow
    # cells survive under generated headers.
    width = len(header_cells)
    for row in body:
        width = max(width, len(row))

    raw_headers = [
        header_cells[i] if i < len(header_cells) else f"col{i + 1}" for i in range(width)
    ]
    headers = _uniquify(raw_headers)

    out: list[dict[str, Any]] = []
    skipped = 0
    for row in body:
        if not any(cell.strip() for cell in row):
            skipped += 1
            continue
        out.append({headers[i]: (row[i].strip() if i < len(row) else "") for i in range(width)})
    return out, skipped, headers


def _finalize(
    all_rows: list[dict[str, Any]],
    skipped: int,
    *,
    sheet_name: str | None,
    sheet_count: int,
) -> ParsedInventory:
    truncated = len(all_rows) > MAX_ROWS
    rows = all_rows[:MAX_ROWS] if truncated else list(all_rows)
    parsed_count = len(rows)
    if truncated:
        rows.append({SENTINEL_KEY: True, "__hint__": f"Input had > {MAX_ROWS} rows."})
    return ParsedInventory(
        rows=rows,
        sheet_name=sheet_name,
        sheet_count=sheet_count,
        rows_parsed=parsed_count,
        rows_skipped=skipped,
        truncated=truncated,
    )


def parse_inventory_detailed(data: bytes, mime_type: str) -> ParsedInventory:
    """Parse `data` into rows plus sheet/row metadata (C-3).

    ``.rows`` holds at most MAX_ROWS data rows, followed by a sentinel
    ``{"__truncated__": True}`` marker when the input was longer.
    """
    kind = kind_for_mime(mime_type)
    if kind == "csv":
        return _parse_csv(data)
    if kind == "xlsx":
        return _parse_xlsx(data)
    raise UnsupportedInventoryFormat(f"Unknown internal kind {kind!r}.")


def parse_inventory(data: bytes, mime_type: str) -> list[dict[str, Any]]:
    """Back-compat: parse `data` into a list of row-dicts (with sentinel).

    Prefer ``parse_inventory_detailed`` when you need the sheet/row metadata.
    """
    return parse_inventory_detailed(data, mime_type).rows


def _parse_csv(data: bytes) -> ParsedInventory:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    grid = [["" if cell is None else str(cell) for cell in row] for row in reader]
    rows, skipped, _ = _grid_to_rows(grid)
    return _finalize(rows, skipped, sheet_name=None, sheet_count=1)


def _parse_xlsx(data: bytes) -> ParsedInventory:
    # openpyxl is lazy-imported so test runs that don't touch XLSX don't
    # pay the import cost.
    from zipfile import BadZipFile

    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException) as exc:
        # Legacy OLE2 .xls bytes or a truncated/corrupt .xlsx. Convert the
        # raw library error into a typed one so the route returns a 422, not
        # an unhandled 500.
        raise CorruptInventoryError(
            "This file could not be read as a valid .xlsx workbook. "
            "Re-save it as .xlsx and upload again."
        ) from exc

    sheet_names = list(wb.sheetnames)
    sheet_count = len(sheet_names)

    # FIX C-3: scan EVERY worksheet and pick the best candidate (most data
    # rows under a plausible header) instead of blindly taking wb.active,
    # which extracts garbage from a workbook that opens on a cover page.
    best: tuple[list[dict[str, Any]], int, str] | None = None
    for name in sheet_names:
        ws = wb[name]
        grid: list[list[str]] = []
        for raw in ws.iter_rows(values_only=True):
            if raw is None:
                grid.append([])
                continue
            grid.append(["" if v is None else str(v).strip() for v in raw])
        rows, skipped, _ = _grid_to_rows(grid)
        if best is None or len(rows) > len(best[0]):
            best = (rows, skipped, name)

    if best is None:
        return _finalize([], 0, sheet_name=None, sheet_count=sheet_count)

    rows, skipped, sheet_name = best
    return _finalize(rows, skipped, sheet_name=sheet_name, sheet_count=sheet_count)
