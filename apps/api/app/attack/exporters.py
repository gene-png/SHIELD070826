"""ATT&CK Coverage deliverable renderers - PDF + XLSX.

XLSX sheets:
  - Heatmap Summary: tactic rollup (counts + coverage %)
  - Coverage:        per-technique status + notes (all 600+ rows)
  - Gaps:            techniques flagged as Gap, ordered by tactic

PDF:
  Executive page with overall coverage % + per-tactic table, then the
  Gap list (top 50 entries).
"""

from __future__ import annotations

import io
from collections.abc import Iterable
from dataclasses import dataclass
from typing import TYPE_CHECKING

from app.attack.analytics import CoverageRollup
from app.attack.catalog import TACTICS, TECHNIQUES, technique_by_id
from app.attack.coverage import CoverageStatus, coverage_label
from app.models.attack_assessment import AttackAssessment, AttackCoverage

if TYPE_CHECKING:
    from reportlab.platypus import TableStyle


@dataclass(frozen=True)
class AttackDeliverableContext:
    client_legal_name: str
    service_title: str
    assessment: AttackAssessment
    coverage: list[AttackCoverage]
    rollup: CoverageRollup


def build_context(
    *,
    client_legal_name: str | None,
    service_title: str,
    assessment: AttackAssessment,
    coverage: Iterable[AttackCoverage],
    rollup: CoverageRollup,
) -> AttackDeliverableContext:
    return AttackDeliverableContext(
        client_legal_name=client_legal_name or "Client",
        service_title=service_title,
        assessment=assessment,
        coverage=list(coverage),
        rollup=rollup,
    )


def _status_or_unscored(value: str | None) -> str:
    if value is None:
        return "Unscored"
    try:
        return coverage_label(CoverageStatus(value))
    except ValueError:
        return "Unknown"


def _tactic_name(tactic_id: str) -> str:
    for t in TACTICS:
        if t.id == tactic_id:
            return t.name
    return tactic_id


# --- Remediation priority ordering (B-7) --------------------------------------
# The old narrative "top remediation gaps" list filtered to GAP, sorted
# ALPHABETICALLY by technique code, and cut at 50 -- alphabetical order is not a
# priority by any measure. We now surface both GAP and PARTIAL techniques (both
# need work) ordered by a defensible priority: the weakest tactic coverage a
# technique touches first, then GAP before PARTIAL, then code as a stable
# tie-break. The section title states the rule so the document explains itself.
_REMEDIATION_STATUSES = (CoverageStatus.GAP.value, CoverageStatus.PARTIAL.value)
_STATUS_RANK = {CoverageStatus.GAP.value: 0, CoverageStatus.PARTIAL.value: 1}
_REMEDIATION_LIMIT = 50
REMEDIATION_SECTION_TITLE = (
    "Remediation priorities (weakest tactic coverage first, then gaps before partials)"
)


def _tactic_coverage_pct(ctx: AttackDeliverableContext) -> dict[str, float]:
    return {tc.tactic_id: tc.coverage_pct for tc in ctx.rollup.by_tactic}


def _remediation_sort_key(cov: AttackCoverage, tactic_cov: dict[str, float]) -> tuple:
    try:
        tactics = technique_by_id(cov.technique_code).tactics
    except KeyError:
        tactics = ()
    # Worst (lowest) coverage among the technique's tactics; an unknown technique
    # (not in the catalog) has no tactics and sorts to the very end.
    worst = min((tactic_cov.get(t, 100.0) for t in tactics), default=999.0)
    return (worst, _STATUS_RANK.get(cov.status or "", 99), cov.technique_code)


def _remediation_rows(
    ctx: AttackDeliverableContext, *, limit: int = _REMEDIATION_LIMIT
) -> list[AttackCoverage]:
    tactic_cov = _tactic_coverage_pct(ctx)
    rows = [c for c in ctx.coverage if c.status in _REMEDIATION_STATUSES]
    rows.sort(key=lambda c: _remediation_sort_key(c, tactic_cov))
    return rows[:limit]


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def render_xlsx(ctx: AttackDeliverableContext) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    header_fill = PatternFill(start_color="FFEEF2F7", end_color="FFEEF2F7", fill_type="solid")
    bold = Font(bold=True)
    italic = Font(italic=True)

    # --- Heatmap Summary ---
    ws = wb.create_sheet("Heatmap Summary")
    ws.append(["Engagement", ctx.client_legal_name])
    ws.append(["Service", ctx.service_title])
    ws.append(["Assessment version", ctx.assessment.version])
    ws.append(["Coverage %", ctx.rollup.coverage_pct])
    ws.append(
        [
            "Scored / Total",
            f"{ctx.rollup.scored_count}/{ctx.rollup.scored_count + ctx.rollup.unscored_count}",
        ]
    )
    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold
    ws.append([])
    headers = [
        "Tactic",
        "Name",
        "Techniques",
        "Sub-techniques",
        "Covered",
        "Partial",
        "Gap",
        "N/A",
        "Unscored",
        "Coverage %",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = bold
        cell.fill = header_fill
    for tc in ctx.rollup.by_tactic:
        ws.append(
            [
                tc.tactic_id,
                tc.tactic_name,
                tc.technique_count,
                tc.sub_technique_count,
                tc.covered,
                tc.partial,
                tc.gap,
                tc.not_applicable,
                tc.unscored,
                tc.coverage_pct,
            ]
        )
    widths = [10, 28, 12, 14, 10, 10, 8, 8, 12, 14]
    for w, col in zip(widths, range(1, len(widths) + 1), strict=True):
        ws.column_dimensions[get_column_letter(col)].width = w

    # --- Coverage (per-technique) ---
    ws2 = wb.create_sheet("Coverage")
    headers2 = ["Technique", "Name", "Tactic(s)", "Type", "Status", "Notes"]
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    cov_by_code = {c.technique_code: c for c in ctx.coverage}
    for tech in TECHNIQUES:
        cov = cov_by_code.get(tech.id)
        tactic_str = ", ".join(_tactic_name(t) for t in tech.tactics)
        ws2.append(
            [
                tech.id,
                tech.name,
                tactic_str,
                "sub" if tech.is_sub_technique else "parent",
                _status_or_unscored(cov.status if cov else None),
                (cov.notes if cov and cov.notes else "") or "",
            ]
        )
    widths2 = [14, 38, 28, 8, 12, 60]
    for w, col in zip(widths2, range(1, len(widths2) + 1), strict=True):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # --- Gaps ---
    ws3 = wb.create_sheet("Gaps")
    headers3 = ["Technique", "Name", "Tactic(s)", "Notes"]
    ws3.append(headers3)
    for col in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=col)
        cell.font = bold
        cell.fill = header_fill
    gap_rows = [c for c in ctx.coverage if c.status == CoverageStatus.GAP.value]
    gap_rows.sort(key=lambda c: c.technique_code)
    for cov in gap_rows:
        try:
            tech = technique_by_id(cov.technique_code)
            tactic_str = ", ".join(_tactic_name(t) for t in tech.tactics)
            name = tech.name
        except KeyError:
            tactic_str = ""
            name = cov.technique_code
        ws3.append([cov.technique_code, name, tactic_str, cov.notes or ""])
    if not gap_rows:
        ws3.append(["—", "No gaps recorded", "", ""])
        ws3.cell(row=2, column=2).font = italic
    widths3 = [14, 38, 28, 60]
    for w, col in zip(widths3, range(1, len(widths3) + 1), strict=True):
        ws3.column_dimensions[get_column_letter(col)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def render_docx(ctx: AttackDeliverableContext) -> bytes:
    """Word deliverable mirroring the PDF (Work Order C4)."""
    from app.docx_export import (
        add_heading,
        add_paragraphs,
        add_table,
        add_title,
        new_document,
        to_bytes,
    )

    doc = new_document(f"{ctx.service_title} — {ctx.client_legal_name}")
    add_title(doc, ctx.service_title, ctx.client_legal_name)

    add_heading(doc, "Coverage summary")
    add_paragraphs(
        doc,
        [
            f"Overall coverage: {ctx.rollup.coverage_pct}%",
            f"Scored: {ctx.rollup.scored_count}/"
            f"{ctx.rollup.scored_count + ctx.rollup.unscored_count}",
            f"Covered {ctx.rollup.covered}, Partial {ctx.rollup.partial}, "
            f"Gap {ctx.rollup.gap}, N/A {ctx.rollup.not_applicable}",
        ],
    )

    add_heading(doc, "Per-tactic rollup")
    add_table(
        doc,
        ["Tactic", "Name", "Covered", "Partial", "Gap", "N/A", "Coverage %"],
        [
            [
                tc.tactic_id,
                tc.tactic_name,
                tc.covered,
                tc.partial,
                tc.gap,
                tc.not_applicable,
                f"{tc.coverage_pct}%",
            ]
            for tc in ctx.rollup.by_tactic
        ],
    )

    remediation = _remediation_rows(ctx)
    add_heading(doc, f"{REMEDIATION_SECTION_TITLE} — top {len(remediation)}")
    if not remediation:
        add_paragraphs(doc, ["No techniques flagged as Gap or Partial."])
    else:
        rows = []
        for cov in remediation:
            try:
                name = technique_by_id(cov.technique_code).name
            except KeyError:
                # B-7: fall back to the code (the PDF already did) so Word never
                # renders a blank technique cell.
                name = cov.technique_code
            rows.append([cov.technique_code, name, _status_or_unscored(cov.status)])
        add_table(doc, ["Code", "Technique", "Status"], rows)

    return to_bytes(doc)


def render_pdf(ctx: AttackDeliverableContext) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
    )

    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.7 * inch,
        bottomMargin=0.7 * inch,
        title=f"{ctx.service_title} — {ctx.client_legal_name}",
        author="SHIELD by Kentro",
    )
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)
    body = styles["BodyText"]

    story: list = []
    story.append(Paragraph(ctx.service_title, h1))
    story.append(Paragraph(ctx.client_legal_name, body))
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("Coverage summary", h2))
    story.append(
        Paragraph(
            f"Overall coverage: <b>{ctx.rollup.coverage_pct}%</b> · "
            f"Scored: <b>{ctx.rollup.scored_count}/"
            f"{ctx.rollup.scored_count + ctx.rollup.unscored_count}</b> · "
            f"Covered <b>{ctx.rollup.covered}</b>, "
            f"Partial <b>{ctx.rollup.partial}</b>, "
            f"Gap <b>{ctx.rollup.gap}</b>, "
            f"N/A <b>{ctx.rollup.not_applicable}</b>",
            body,
        )
    )

    story.append(Paragraph("Per-tactic rollup", h2))
    tactic_table_data: list[list] = [
        ["Tactic", "Name", "Covered", "Partial", "Gap", "N/A", "Coverage %"]
    ]
    for tc in ctx.rollup.by_tactic:
        tactic_table_data.append(
            [
                tc.tactic_id,
                tc.tactic_name,
                tc.covered,
                tc.partial,
                tc.gap,
                tc.not_applicable,
                f"{tc.coverage_pct}%",
            ]
        )
    tactic_col_widths = [
        0.8 * inch,
        1.9 * inch,
        0.7 * inch,
        0.7 * inch,
        0.6 * inch,
        0.6 * inch,
        0.9 * inch,
    ]
    tactic_table = Table(
        tactic_table_data,
        colWidths=tactic_col_widths,
        repeatRows=1,
    )
    tactic_table.setStyle(_table_style())
    story.append(tactic_table)

    story.append(PageBreak())

    # Prioritized remediation list (B-7): weakest tactic coverage first, then
    # gaps before partials; not an alphabetical dump.
    remediation = _remediation_rows(ctx)
    story.append(Paragraph(f"{REMEDIATION_SECTION_TITLE} — top {len(remediation)}", h2))
    if not remediation:
        story.append(Paragraph("No techniques flagged as Gap or Partial.", body))
    else:
        gap_table_data: list[list] = [["Code", "Technique", "Status"]]
        for cov in remediation:
            try:
                name = technique_by_id(cov.technique_code).name
            except KeyError:
                name = cov.technique_code
            gap_table_data.append([cov.technique_code, name, _status_or_unscored(cov.status)])
        gap_table = Table(
            gap_table_data,
            colWidths=[1.1 * inch, 3.8 * inch, 0.8 * inch],
            repeatRows=1,
        )
        gap_table.setStyle(_table_style())
        story.append(gap_table)

    doc.build(story)
    return out.getvalue()


def _table_style() -> TableStyle:
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle

    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f7")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0e1220")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d6dae3")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ]
    )


__all__ = ["AttackDeliverableContext", "build_context", "render_pdf", "render_xlsx"]
