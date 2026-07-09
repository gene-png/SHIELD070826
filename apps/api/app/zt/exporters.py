"""Zero Trust deliverable renderers - PDF + XLSX from a ZT assessment.

Three-sheet XLSX (Score Summary / Answers / Gap Plan) + executive PDF
(overall stage + per-pillar table, then top-N gap table). Pure
functions, no I/O. Framework-aware: CISA labels render Traditional/
Initial/..., DoD labels render Baseline/Target/...
"""

from __future__ import annotations

import io
from collections.abc import Iterable
from dataclasses import dataclass
from typing import TYPE_CHECKING

from app.models.zt_assessment import ZtAnswer, ZtAssessment
from app.zt.catalog import capabilities, pillars
from app.zt.maturity import ZtFrameworkCode, stage_label
from app.zt.scoring import GapAnalysis, ScoreResult

if TYPE_CHECKING:
    from reportlab.platypus import TableStyle


@dataclass(frozen=True)
class ZtDeliverableContext:
    client_legal_name: str
    service_title: str
    framework: ZtFrameworkCode
    assessment: ZtAssessment
    answers: list[ZtAnswer]
    score: ScoreResult
    gap: GapAnalysis


def build_context(
    *,
    client_legal_name: str | None,
    service_title: str,
    framework: ZtFrameworkCode,
    assessment: ZtAssessment,
    answers: Iterable[ZtAnswer],
    score: ScoreResult,
    gap: GapAnalysis,
) -> ZtDeliverableContext:
    return ZtDeliverableContext(
        client_legal_name=client_legal_name or "Client",
        service_title=service_title,
        framework=framework,
        assessment=assessment,
        answers=list(answers),
        score=score,
        gap=gap,
    )


def _fmt(value: float | None) -> str:
    if value is None:
        return "—"
    return f"{value:.2f}"


def _framework_label(framework: ZtFrameworkCode) -> str:
    return (
        "CISA ZTMM 2.0"
        if framework == ZtFrameworkCode.CISA_ZTMM_2_0
        else "DoD ZT Reference Architecture"
    )


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def render_xlsx(ctx: ZtDeliverableContext) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    header_fill = PatternFill(start_color="FFEEF2F7", end_color="FFEEF2F7", fill_type="solid")
    bold = Font(bold=True)
    italic = Font(italic=True)

    # --- Score Summary ---
    ws = wb.create_sheet("Score Summary")
    ws.append(["Engagement", ctx.client_legal_name])
    ws.append(["Service", ctx.service_title])
    ws.append(["Framework", _framework_label(ctx.framework)])
    ws.append(["Assessment version", ctx.assessment.version])
    ws.append(["Overall stage", ctx.score.overall_stage_label])
    ws.append(["Average stage", _fmt(ctx.score.average_stage)])
    ws.append(["Coverage", f"{ctx.score.answered_capabilities}/{ctx.score.total_capabilities}"])
    for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold
    ws.append([])
    ws.append(["Pillar", "Name", "Answered", "Total", "Coverage %", "Average stage"])
    for col_idx in range(1, 7):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = bold
        cell.fill = header_fill
    for ps in ctx.score.by_pillar:
        ws.append(
            [
                ps.pillar_code,
                ps.pillar_name,
                ps.answered_count,
                ps.capability_count,
                ps.coverage_pct,
                _fmt(ps.average_stage),
            ]
        )
    for w, col in zip([10, 36, 12, 10, 14, 16], range(1, 7), strict=True):
        ws.column_dimensions[get_column_letter(col)].width = w

    # --- Answers ---
    ws2 = wb.create_sheet("Answers")
    headers = ["Capability", "Pillar", "Name", "Outcome", "Stage", "Stage label", "Notes"]
    ws2.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    answers_by_code = {a.capability_code: a for a in ctx.answers}
    # Use catalog order so missing answers still render as blank rows.
    pillar_lookup = {p.code: p.name for p in pillars(ctx.framework)}
    for cap in capabilities(ctx.framework):
        ans = answers_by_code.get(cap.code)
        s = ans.maturity_stage if ans else None
        notes = ans.notes if ans else None
        ws2.append(
            [
                cap.code,
                f"{cap.pillar_code} · {pillar_lookup.get(cap.pillar_code, cap.pillar_code)}",
                cap.name,
                cap.outcome,
                s if s is not None else "",
                stage_label(s, ctx.framework) if s is not None else "Unscored",
                notes or "",
            ]
        )
    for w, col in zip([18, 30, 36, 60, 8, 16, 60], range(1, 8), strict=True):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # --- Gap Plan ---
    ws3 = wb.create_sheet("Gap Plan")
    headers3 = [
        "Capability",
        "Pillar",
        "Name",
        "Current stage",
        "Target stage",
        "Gap size",
        "Priority",
        "Notes",
    ]
    ws3.append(headers3)
    for col in range(1, len(headers3) + 1):
        cell = ws3.cell(row=1, column=col)
        cell.font = bold
        cell.fill = header_fill
    for g in ctx.gap.gaps:
        ws3.append(
            [
                g.code,
                g.pillar_code,
                g.name,
                g.current_stage,
                g.target_stage,
                g.gap_size,
                g.priority_score,
                g.notes or "",
            ]
        )
    if not ctx.gap.gaps:
        ws3.append(["—", "", "No gaps at target stage", "", ctx.gap.target_stage, 0, 0, ""])
        ws3.cell(row=2, column=3).font = italic
    for w, col in zip([18, 10, 36, 14, 14, 12, 12, 50], range(1, 9), strict=True):
        ws3.column_dimensions[get_column_letter(col)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# HTML dashboard (primary, auto-built, in-browser view)
# ---------------------------------------------------------------------------


def render_html_dashboard(ctx: ZtDeliverableContext) -> bytes:
    """Self-contained Zero Trust maturity dashboard (dark theme, no external
    assets). Mirrors the Tech Debt dashboard: KPI cards, per-pillar maturity
    bars, a prioritized remediation roadmap, then the full answer set. Every
    dynamic string is HTML-escaped."""
    from html import escape

    from app.zt.maturity import level_count, stage_label

    max_stage = level_count(ctx.framework)
    score = ctx.score
    gap = ctx.gap
    initials = "".join(w[0] for w in ctx.client_legal_name.split()[:2]).upper() or "ZT"

    def _stage_tone(avg: float | None) -> str:
        if avg is None:
            return "muted"
        ratio = avg / max_stage if max_stage else 0
        if ratio >= 0.75:
            return "green"
        if ratio >= 0.5:
            return "amber"
        return "red"

    # ---- KPI cards ----
    avg = score.average_stage
    kpis = f"""
    <div class="kpi">
      <div class="label">Overall Maturity</div>
      <div class="value">{escape(score.overall_stage_label)}</div>
      <div class="sub">{_framework_label(ctx.framework)}</div>
    </div>
    <div class="kpi amber">
      <div class="label">Average Stage</div>
      <div class="value">{_fmt(avg)}<span style="font-size:16px;color:var(--muted)"> / {max_stage}</span></div>
      <div class="sub">Across {len(score.by_pillar)} pillars</div>
    </div>
    <div class="kpi">
      <div class="label">Coverage</div>
      <div class="value">{score.coverage_pct}%</div>
      <div class="sub">{score.answered_capabilities} of {score.total_capabilities} capabilities scored</div>
    </div>
    <div class="kpi red">
      <div class="label">Gaps to Close</div>
      <div class="value">{gap.total_gap_count}</div>
      <div class="sub">To reach target {escape(gap.target_label)} (S{gap.target_stage})</div>
    </div>
    """

    # ---- Per-pillar maturity bars ----
    pillar_bars = "".join(f"""
      <div class="bar-row">
        <div class="bar-label" title="{escape(ps.pillar_name)}">{escape(ps.pillar_code)} · {escape(ps.pillar_name)}</div>
        <div class="bar-track"><div class="bar-fill {_stage_tone(ps.average_stage)}" style="width:{((ps.average_stage or 0) / max_stage * 100) if max_stage else 0:.1f}%"></div></div>
        <div class="bar-value">{_fmt(ps.average_stage)} / {max_stage} · {ps.coverage_pct}%</div>
      </div>""" for ps in score.by_pillar)

    # ---- Remediation roadmap ----
    if gap.gaps:
        roadmap = "".join(f"""
        <div class="rcard {'high' if g.gap_size >= 2 else 'medium'}">
          <div class="rhead">
            <span class="cat">{escape(g.name)}</span>
            <span class="sev">P{g.priority_score:.1f}</span>
          </div>
          <div class="reco"><b>{escape(g.pillar_code)}</b> · {escape(g.code)} — move from
            <b>S{g.current_stage} ({escape(stage_label(g.current_stage, ctx.framework))})</b> to
            <b>S{g.target_stage} ({escape(stage_label(g.target_stage, ctx.framework))})</b>.
            {escape(g.notes or '')}</div>
        </div>""" for g in gap.gaps)
    else:
        roadmap = f'<p class="desc">No gaps at target stage S{gap.target_stage} ({escape(gap.target_label)}). Every scored capability is at or above target.</p>'

    # ---- Full answers table ----
    from app.zt.catalog import capabilities, pillars

    answers_by_code = {a.capability_code: a for a in ctx.answers}
    pillar_lookup = {p.code: p.name for p in pillars(ctx.framework)}
    rows = ""
    for cap in capabilities(ctx.framework):
        ans = answers_by_code.get(cap.code)
        s = ans.maturity_stage if ans else None
        notes = (ans.notes if ans else None) or ""
        label = stage_label(s, ctx.framework) if s is not None else "Unscored"
        tone = _stage_tone(float(s)) if s is not None else "muted"
        rows += f"""
        <tr>
          <td>{escape(cap.code)}</td>
          <td><span class="cat-chip">{escape(cap.pillar_code)}</span> {escape(pillar_lookup.get(cap.pillar_code, ''))}</td>
          <td>{escape(cap.name)}</td>
          <td><span class="disp {tone}">{('S' + str(s) + ' · ') if s is not None else ''}{escape(label)}</span></td>
          <td>{escape(notes)}</td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1.0" />
<title>{escape(ctx.client_legal_name)} — {escape(ctx.service_title)}</title>
<style>
  :root {{
    --bg:#0b1020; --bg-2:#11172d; --panel:#151b35; --panel-2:#1a2143; --border:#232a4d;
    --text:#e6e9f5; --muted:#98a2c4; --accent:#6366f1; --accent-2:#22d3ee;
    --green:#10b981; --amber:#f59e0b; --red:#ef4444;
  }}
  * {{ box-sizing:border-box; }}
  html,body {{ margin:0; padding:0; background:
      radial-gradient(1200px 600px at 0% 0%, #1a2046 0%, var(--bg) 50%) fixed,
      radial-gradient(900px 600px at 100% 0%, #1a1d3d 0%, transparent 60%) fixed, var(--bg);
    color:var(--text); font-family:-apple-system,BlinkMacSystemFont,"Inter","Segoe UI",Roboto,sans-serif;
    min-height:100vh; -webkit-font-smoothing:antialiased; }}
  .wrap {{ max-width:1440px; margin:0 auto; padding:32px 28px 64px; }}
  .header {{ display:flex; align-items:center; justify-content:space-between; margin-bottom:28px; flex-wrap:wrap; gap:16px; }}
  .brand {{ display:flex; align-items:center; gap:14px; }}
  .logo {{ width:44px; height:44px; border-radius:12px; background:linear-gradient(135deg,var(--accent),var(--accent-2));
    display:flex; align-items:center; justify-content:center; font-weight:800; color:#fff; font-size:16px; box-shadow:0 8px 24px rgba(99,102,241,.35); }}
  .brand h1 {{ margin:0; font-size:20px; font-weight:700; letter-spacing:-.01em; }}
  .brand p {{ margin:2px 0 0; font-size:13px; color:var(--muted); }}
  .badge {{ display:inline-flex; align-items:center; gap:6px; background:rgba(99,102,241,.12); color:#c7d2fe;
    border:1px solid rgba(99,102,241,.35); padding:6px 12px; border-radius:999px; font-size:12px; font-weight:600; }}
  .badge .dot {{ width:8px; height:8px; border-radius:50%; background:var(--green); box-shadow:0 0 0 4px rgba(16,185,129,.18); }}
  .kpis {{ display:grid; grid-template-columns:repeat(4,1fr); gap:16px; margin-bottom:22px; }}
  .kpi {{ background:linear-gradient(180deg,var(--panel) 0%,var(--panel-2) 100%); border:1px solid var(--border);
    border-radius:16px; padding:20px 22px; position:relative; overflow:hidden; }}
  .kpi::after {{ content:""; position:absolute; right:-30px; top:-30px; width:120px; height:120px; border-radius:50%;
    opacity:.14; background:radial-gradient(circle,var(--accent),transparent 70%); }}
  .kpi.amber::after {{ background:radial-gradient(circle,var(--amber),transparent 70%); }}
  .kpi.red::after {{ background:radial-gradient(circle,var(--red),transparent 70%); }}
  .kpi .label {{ font-size:12px; color:var(--muted); text-transform:uppercase; letter-spacing:.08em; font-weight:600; }}
  .kpi .value {{ font-size:30px; font-weight:700; margin-top:8px; letter-spacing:-.02em; }}
  .kpi .sub {{ font-size:12px; color:var(--muted); margin-top:6px; }}
  .section {{ background:linear-gradient(180deg,var(--panel) 0%,var(--panel-2) 100%); border:1px solid var(--border);
    border-radius:16px; padding:22px; margin-bottom:18px; }}
  .section h2 {{ margin:0 0 4px; font-size:16px; font-weight:700; display:flex; align-items:center; gap:10px; }}
  .section h2 .pill {{ font-size:11px; padding:3px 8px; border-radius:999px; background:rgba(99,102,241,.18); color:#c7d2fe; font-weight:600; }}
  .section .desc {{ color:var(--muted); font-size:13px; margin:0 0 16px; }}
  .bar-row {{ display:grid; grid-template-columns:260px 1fr 150px; align-items:center; gap:12px; margin-bottom:10px; }}
  .bar-label {{ font-size:12.5px; color:var(--text); white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
  .bar-track {{ height:14px; background:var(--bg-2); border-radius:999px; overflow:hidden; }}
  .bar-fill {{ height:100%; border-radius:999px; background:linear-gradient(90deg,var(--accent),var(--accent-2)); }}
  .bar-fill.green {{ background:linear-gradient(90deg,var(--green),var(--accent-2)); }}
  .bar-fill.amber {{ background:linear-gradient(90deg,var(--amber),var(--accent)); }}
  .bar-fill.red {{ background:linear-gradient(90deg,var(--red),var(--amber)); }}
  .bar-fill.muted {{ background:var(--border); }}
  .bar-value {{ font-size:12px; color:var(--muted); text-align:right; font-variant-numeric:tabular-nums; }}
  .redundancy-grid {{ display:grid; grid-template-columns:repeat(2,1fr); gap:14px; }}
  .rcard {{ background:var(--bg-2); border:1px solid var(--border); border-left:3px solid var(--amber); border-radius:12px;
    padding:16px 18px; display:flex; flex-direction:column; gap:8px; }}
  .rcard.high {{ border-left-color:var(--red); }}
  .rcard.medium {{ border-left-color:var(--amber); }}
  .rcard .rhead {{ display:flex; align-items:center; justify-content:space-between; }}
  .rcard .cat {{ font-weight:700; font-size:14px; }}
  .rcard .sev {{ font-size:10px; font-weight:700; text-transform:uppercase; letter-spacing:.08em; padding:3px 8px; border-radius:999px;
    background:rgba(99,102,241,.18); color:#c7d2fe; }}
  .reco {{ font-size:12.5px; color:var(--muted); line-height:1.55; }}
  .reco b {{ color:var(--text); font-weight:600; }}
  .table-scroll {{ max-height:560px; overflow-y:auto; border:1px solid var(--border); border-radius:10px; }}
  table {{ width:100%; border-collapse:collapse; font-size:12.5px; }}
  thead th {{ position:sticky; top:0; background:var(--bg-2); color:var(--muted); text-align:left; padding:10px 12px;
    font-weight:600; text-transform:uppercase; letter-spacing:.06em; font-size:11px; border-bottom:1px solid var(--border); }}
  tbody td {{ padding:10px 12px; border-bottom:1px solid rgba(255,255,255,.04); vertical-align:top; }}
  tbody tr:hover {{ background:rgba(99,102,241,.04); }}
  .cat-chip {{ display:inline-block; padding:2px 8px; border-radius:999px; font-size:11px; font-weight:600; background:rgba(99,102,241,.15); color:#c7d2fe; }}
  .disp {{ font-size:11px; font-weight:700; padding:2px 8px; border-radius:999px; }}
  .disp.green {{ background:rgba(16,185,129,.15); color:#a7f3d0; }}
  .disp.amber {{ background:rgba(245,158,11,.15); color:#fde68a; }}
  .disp.red {{ background:rgba(239,68,68,.15); color:#fecaca; }}
  .disp.muted {{ background:rgba(152,162,196,.15); color:var(--muted); }}
  footer {{ color:var(--muted); font-size:11.5px; text-align:center; margin-top:24px; }}
  @media (max-width:1080px) {{ .kpis{{grid-template-columns:repeat(2,1fr);}} .redundancy-grid{{grid-template-columns:1fr;}}
    .bar-row{{grid-template-columns:150px 1fr 110px;}} }}
</style>
</head>
<body>
<div class="wrap">
  <div class="header">
    <div class="brand">
      <div class="logo">{escape(initials)}</div>
      <div>
        <h1>{escape(ctx.client_legal_name)} — Zero Trust Maturity</h1>
        <p>Executive Dashboard · {escape(_framework_label(ctx.framework))}</p>
      </div>
    </div>
    <div class="badge"><span class="dot"></span> SHIELD by Kentro · Assessment v{ctx.assessment.version}</div>
  </div>

  <div class="kpis">{kpis}</div>

  <div class="section">
    <h2>Maturity by Pillar <span class="pill">{len(score.by_pillar)} pillars</span></h2>
    <p class="desc">Average maturity stage per pillar (out of {max_stage}). Coverage shows how much of the pillar was scored.</p>
    {pillar_bars}
  </div>

  <div class="section">
    <h2>Remediation Roadmap <span class="pill">{gap.total_gap_count} gaps</span></h2>
    <p class="desc">Prioritized capabilities below the target stage S{gap.target_stage} ({escape(gap.target_label)}). Higher P-score = higher priority.</p>
    <div class="redundancy-grid">{roadmap}</div>
  </div>

  <div class="section">
    <h2>Full Assessment <span class="pill">{score.total_capabilities} capabilities</span></h2>
    <p class="desc">Every capability in the framework and its scored maturity stage.</p>
    <div class="table-scroll">
      <table>
        <thead><tr><th>Code</th><th>Pillar</th><th>Capability</th><th>Stage</th><th>Notes</th></tr></thead>
        <tbody>{rows}</tbody>
      </table>
    </div>
  </div>

  <footer>Generated by SHIELD by Kentro · {escape(ctx.service_title)} · Overall {escape(score.overall_stage_label)} · {score.coverage_pct}% coverage</footer>
</div>
</body>
</html>""".encode()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def render_docx(ctx: ZtDeliverableContext) -> bytes:
    """Word deliverable mirroring the PDF (Work Order C4)."""
    from app.docx_export import (
        add_heading,
        add_paragraphs,
        add_table,
        add_title,
        new_document,
        to_bytes,
    )

    doc = new_document(f"{ctx.service_title} — {ctx.client_legal_name}")
    add_title(
        doc,
        ctx.service_title,
        f"{ctx.client_legal_name} · {_framework_label(ctx.framework)}",
    )

    add_heading(doc, "Maturity summary")
    add_paragraphs(
        doc,
        [
            f"Overall stage: {ctx.score.overall_stage_label}",
            f"Average stage: {_fmt(ctx.score.average_stage)}",
            f"Coverage: {ctx.score.answered_capabilities}/"
            f"{ctx.score.total_capabilities} ({ctx.score.coverage_pct}%)",
        ],
    )

    add_heading(doc, "Per-pillar rollup")
    add_table(
        doc,
        ["Pillar", "Name", "Average stage", "Coverage"],
        [
            [
                ps.pillar_code,
                ps.pillar_name,
                _fmt(ps.average_stage),
                f"{ps.answered_count}/{ps.capability_count} ({ps.coverage_pct}%)",
            ]
            for ps in ctx.score.by_pillar
        ],
    )

    add_heading(doc, f"Top remediation gaps (target S{ctx.gap.target_stage})")
    if not ctx.gap.gaps:
        add_paragraphs(
            doc,
            [f"No gaps at target stage {ctx.gap.target_stage} " f"({ctx.gap.target_label})."],
        )
    else:
        add_table(
            doc,
            ["Code", "Pillar", "Capability", "Current → Target", "Priority"],
            [
                [
                    g.code,
                    g.pillar_code,
                    g.name,
                    f"S{g.current_stage} → S{g.target_stage}",
                    f"{g.priority_score:.2f}",
                ]
                for g in ctx.gap.gaps
            ],
        )

    return to_bytes(doc)


def render_pdf(ctx: ZtDeliverableContext) -> bytes:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
    )

    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.7 * inch,
        bottomMargin=0.7 * inch,
        title=f"{ctx.service_title} — {ctx.client_legal_name}",
        author="SHIELD by Kentro",
    )
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)
    body = styles["BodyText"]

    story: list = []
    story.append(Paragraph(ctx.service_title, h1))
    story.append(Paragraph(f"{ctx.client_legal_name} · {_framework_label(ctx.framework)}", body))
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("Maturity summary", h2))
    story.append(
        Paragraph(
            f"Overall stage: <b>{ctx.score.overall_stage_label}</b> · "
            f"Average stage: <b>{_fmt(ctx.score.average_stage)}</b> · "
            f"Coverage: <b>{ctx.score.answered_capabilities}/"
            f"{ctx.score.total_capabilities}</b> "
            f"({ctx.score.coverage_pct}%)",
            body,
        )
    )

    story.append(Paragraph("Per-pillar rollup", h2))
    fn_table_data: list[list] = [["Pillar", "Name", "Average stage", "Coverage"]]
    for ps in ctx.score.by_pillar:
        fn_table_data.append(
            [
                ps.pillar_code,
                ps.pillar_name,
                _fmt(ps.average_stage),
                f"{ps.answered_count}/{ps.capability_count} ({ps.coverage_pct}%)",
            ]
        )
    fn_table = Table(
        fn_table_data,
        colWidths=[0.8 * inch, 3.2 * inch, 1.2 * inch, 1.6 * inch],
        repeatRows=1,
    )
    fn_table.setStyle(_table_style())
    story.append(fn_table)

    story.append(PageBreak())

    story.append(Paragraph(f"Top remediation gaps (target S{ctx.gap.target_stage})", h2))
    if not ctx.gap.gaps:
        story.append(
            Paragraph(
                f"No gaps at target stage {ctx.gap.target_stage} " f"({ctx.gap.target_label}).",
                body,
            )
        )
    else:
        gap_table_data: list[list] = [
            ["Code", "Pillar", "Capability", "Current → Target", "Priority"]
        ]
        for g in ctx.gap.gaps:
            gap_table_data.append(
                [
                    g.code,
                    g.pillar_code,
                    g.name,
                    f"S{g.current_stage} → S{g.target_stage}",
                    f"{g.priority_score:.2f}",
                ]
            )
        gap_table = Table(
            gap_table_data,
            colWidths=[1.1 * inch, 0.8 * inch, 2.9 * inch, 1.4 * inch, 0.8 * inch],
            repeatRows=1,
        )
        gap_table.setStyle(_table_style())
        story.append(gap_table)

    doc.build(story)
    return out.getvalue()


def _table_style() -> TableStyle:
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle

    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2f7")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0e1220")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d6dae3")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ]
    )


__all__ = [
    "ZtDeliverableContext",
    "build_context",
    "render_docx",
    "render_html_dashboard",
    "render_pdf",
    "render_xlsx",
]
